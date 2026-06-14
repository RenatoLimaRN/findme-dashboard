#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
enriquecer_atividades.py — Enriquece a aba "Atividades" de um relatório FindMe
GERAL com:

  1. Cor de fundo por Status nas linhas de dados reais
     (verde=Completa, amarelo=Parcial/Incompleta,
     vermelho=Perdida/Não iniciada, vermelho-escuro=Esperada-Não-Registrada).
  2. Linhas extras pras avulsas esperadas (postos/*.json) que nem foram
     criadas no relatório FindMe — APENAS para grupos que não têm a seção
     nativa "📋 Atividades configuradas" (essa seção já cobre o caso).
  3. Atualização do resumo no cabeçalho de cada local (📍).

Uso:
    python enriquecer_atividades.py "<arquivo.xlsx>" "<dados.json>"

IDEMPOTENTE: pode ser rodado várias vezes no mesmo arquivo. Antes de aplicar,
remove qualquer enriquecimento anterior (linhas [ESPERADA — NÃO REGISTRADA] e
cores nas linhas de dados) e re-aplica do zero.

Respeita as seções nativas do `findme_programacao.py`:
  - Cabeçalhos `📍 LOCAL` (merge A:K) — re-mergidos no fim pra fix do openpyxl
  - Seção `📋 Atividades configuradas` — NÃO é colorida nem duplicada
"""
import sys
import json
import re
import unicodedata
from datetime import date
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl nao instalado. Rode: pip install openpyxl")
    sys.exit(1)

DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
COL_FINAL = 11  # colunas da aba Atividades vão até K

# Marcadores que identificam linhas/sections
TAG_ESPERADA = "[ESPERADA — NÃO REGISTRADA]"
ICONE_GRUPO = "📍"
ICONE_CONFIG = "📋"

# Status que reconhecemos como linha de dado válida (case-insensitive)
STATUS_VALIDOS = {
    "completa", "incompleta", "incompleta c/ justif.",
    "incompleta c/ justificativa", "perdida", "nao iniciada",
    "não iniciada", "esperada não registrada",
}

# Paleta de cores
COR_OK             = "C6EFCE"
COR_PARCIAL        = "FFEB9C"
COR_PERDIDA        = "FFC7CE"
COR_ESPERADA_FALTA = "E57373"
FILL_BRANCO        = PatternFill(fill_type=None)

# Paleta suave da aba unificada — cor SÓ na célula de Status (fundo, texto).
# O resto da linha fica branco/cinza-claro pra não poluir.
STATUS_CORES = {
    "feita":          ("E6F4EA", "1E6B3C"),
    "parcial":        ("FEF3D7", "7F6000"),
    "nao_feita":      ("FDE8E8", "A12D2D"),
    "nao_registrada": ("F0EEE9", "5F5E5A"),
}
STATUS_ROTULO = {
    "feita": "feita", "parcial": "parcial",
    "nao_feita": "não feita", "nao_registrada": "não registrada",
}
COR_LINHA_SISTEMA = "FFF7ED"   # creme leve — destaca atividade vinda do sistema
COR_ZEBRA         = "FAFAF8"   # cinza-claro alternado nas linhas de cadastro


def _slug(nome):
    s = unicodedata.normalize("NFD", nome or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"^\s*CONDOMINIO\s+", "", s, flags=re.I).strip()
    return re.sub(r"[^A-Za-z0-9]+", "_", s.lower()).strip("_")


def _data_br(yyyy_mm_dd):
    try:
        return date.fromisoformat(yyyy_mm_dd).strftime("%d/%m/%Y")
    except Exception:
        return None


def _norm_modelo(m):
    s = unicodedata.normalize("NFD", m or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.upper().strip())


def _bucket(status_str):
    s = (status_str or "").strip().lower()
    if s == "completa":
        return "ok"
    if "incompleta" in s:
        return "parcial"
    if "esperada" in s:
        return "esperada_falta"
    if s in ("perdida", "nao iniciada", "não iniciada"):
        return "nao_feita"
    return None  # NÃO é um status conhecido — não colorir


def _cor(bucket):
    return {"ok": COR_OK, "parcial": COR_PARCIAL,
            "nao_feita": COR_PERDIDA, "esperada_falta": COR_ESPERADA_FALTA}.get(bucket)


# ─── Etapa 1: remover enriquecimento anterior (idempotência) ──────────────────

def _remover_enriquecimento_anterior(ws):
    """Apaga linhas [ESPERADA — NÃO REGISTRADA] e limpa fills nas linhas
    de dados reais (preserva fills nativos de cabeçalhos e seções)."""
    # 1a) achar linhas com [ESPERADA — NÃO REGISTRADA] e deletar (de baixo pra cima)
    rows_remove = []
    for ri in range(1, ws.max_row + 1):
        v4 = ws.cell(row=ri, column=4).value
        if v4 and TAG_ESPERADA in str(v4):
            rows_remove.append(ri)
    for ri in sorted(rows_remove, reverse=True):
        ws.delete_rows(ri, amount=1)

    # 1b) limpar fills:
    #   (a) em linhas de dados reais (DATE em col1)
    #   (b) em linhas 📋 "Atividades configuradas" (versão antiga colorava
    #       errado essas)
    #   (c) em qualquer linha que tem nossa cor conhecida (paranoid)
    # Preserva 📍 (cabeçalho de grupo — formatação nativa azul-escura)
    nossas_cores = {
        "00C6EFCE", "C6EFCE", "00FFEB9C", "FFEB9C",
        "00FFC7CE", "FFC7CE", "00E57373", "E57373",
    }
    for ri in range(1, ws.max_row + 1):
        v1 = ws.cell(row=ri, column=1).value
        v1s = str(v1 or "").strip()
        if ICONE_GRUPO in v1s:
            continue  # preserva formatação nativa do 📍
        is_data = bool(DATE_RE.match(v1s))
        is_config_header = ICONE_CONFIG in v1s or "Atividades configuradas" in v1s
        cor_atual = None
        f = ws.cell(row=ri, column=1).fill
        if f and f.start_color:
            cor_atual = f.start_color.rgb
        is_nossa_cor = cor_atual in nossas_cores
        if is_data or is_config_header or is_nossa_cor:
            for ci in range(1, COL_FINAL + 1):
                ws.cell(row=ri, column=ci).fill = FILL_BRANCO
    return len(rows_remove)


# ─── Etapa 2: escanear grupos ─────────────────────────────────────────────────

def _escanear_grupos(ws):
    """Identifica os grupos 📍 e detecta se cada um tem seção nativa 📋
    "Atividades configuradas". Retorna lista de dicts."""
    grupos = []
    current = None
    for ri in range(1, ws.max_row + 1):
        v = ws.cell(row=ri, column=1).value
        s = str(v or "").strip()
        if ICONE_GRUPO in s:
            if current is not None:
                grupos.append(current)
            local = re.split(r"[—\-]", s.replace(ICONE_GRUPO, ""), maxsplit=1)[0].strip()
            current = {
                "header_row": ri,
                "nome_local": local,
                "slug_local": _slug(local),
                "data_end_row": ri,
                "tem_secao_configurada": False,
                "modelos_na_secao_config": set(),
            }
        elif current is not None:
            if DATE_RE.match(s):
                current["data_end_row"] = ri
            elif ICONE_CONFIG in s or "Atividades configuradas" in s:
                current["tem_secao_configurada"] = True
            elif current.get("tem_secao_configurada") and not DATE_RE.match(s):
                # dentro da seção configurada — coleta modelos (col3 = Modelo de Atividade)
                modelo = ws.cell(row=ri, column=3).value
                if modelo:
                    current["modelos_na_secao_config"].add(_norm_modelo(modelo))
    if current is not None:
        grupos.append(current)
    return grupos


def _modelos_dados_no_grupo(ws, header_row, end_row):
    modelos = set()
    for ri in range(header_row + 1, end_row + 1):
        v1 = str(ws.cell(row=ri, column=1).value or "").strip()
        if not DATE_RE.match(v1):
            continue
        v4 = str(ws.cell(row=ri, column=4).value or "").strip()
        v4 = re.sub(r"^\[AVULSA\]\s*", "", v4).strip()
        if v4:
            modelos.add(_norm_modelo(v4))
    return modelos


# ─── Etapa 3: aplicar enriquecimento ──────────────────────────────────────────

def enriquecer(xlsx_path, dados):
    wb = openpyxl.load_workbook(xlsx_path)
    if "Atividades" not in wb.sheetnames:
        return {"erro": "aba 'Atividades' nao encontrada"}
    ws = wb["Atividades"]

    cruz = dados.get("cruzamento_por_local", {})
    data_alvo_br = _data_br(dados.get("dia_alvo"))
    # Index para resolver "qual cruzamento usar pra esse grupo" — útil
    # quando o nome do local varia (fuzzy match já foi resolvido no leitor).
    # Aqui só precisamos do cruzamento direto por slug.

    # 1) limpar enriquecimento anterior (idempotência)
    n_removidas = _remover_enriquecimento_anterior(ws)

    # 2) escanear grupos
    grupos = _escanear_grupos(ws)
    if not grupos:
        return {"erro": "nenhum grupo 📍 encontrado na aba Atividades"}

    # 3) determinar quais grupos precisam de [ESPERADA — NÃO REGISTRADA]
    # APENAS pra grupos SEM seção "📋 Atividades configuradas"
    inserir_por_grupo = {}
    for i, g in enumerate(grupos):
        if g["tem_secao_configurada"]:
            continue  # seção nativa já mostra o que falta — não duplicar
        c = cruz.get(g["slug_local"])
        if not c or not c.get("esperadas_detalhe"):
            continue
        prox_header = grupos[i + 1]["header_row"] - 1 if i + 1 < len(grupos) else ws.max_row
        modelos_dados = _modelos_dados_no_grupo(ws, g["header_row"], prox_header)
        faltantes = [e for e in c["esperadas_detalhe"]
                     if _norm_modelo(e["modelo"]) not in modelos_dados]
        if faltantes:
            inserir_por_grupo[i] = faltantes

    # 4) inserir linhas (de baixo pra cima)
    n_inseridas = 0
    for i in sorted(inserir_por_grupo.keys(), reverse=True):
        g = grupos[i]
        faltantes = inserir_por_grupo[i]
        insert_at = g["data_end_row"] + 1
        ws.insert_rows(insert_at, amount=len(faltantes))
        for k, esp in enumerate(faltantes):
            r = insert_at + k
            ws.cell(row=r, column=1, value=data_alvo_br or "—")
            ws.cell(row=r, column=2, value="—")
            ws.cell(row=r, column=3, value="—")
            modelo = esp.get("modelo", "")
            posto = esp.get("posto", "")
            tag = f"{TAG_ESPERADA} {modelo}"
            if posto:
                tag += f"  ({posto})"
            ws.cell(row=r, column=4, value=tag)
            ws.cell(row=r, column=5, value="Esperada não registrada")
        n_inseridas += len(faltantes)

    # 5) Re-aplicar merge A:K em TODAS as linhas com 📍 (fix do openpyxl)
    # E também na seção 📋 "Atividades configuradas" que também merge
    for ri in range(1, ws.max_row + 1):
        v = ws.cell(row=ri, column=1).value
        s = str(v or "")
        if ICONE_GRUPO in s or ICONE_CONFIG in s:
            # remove merges existentes que cubram essa linha
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row <= ri <= mr.max_row:
                    try:
                        ws.unmerge_cells(str(mr))
                    except Exception:
                        pass
            # re-merge a linha inteira
            try:
                ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=COL_FINAL)
            except Exception:
                pass

    # 6) Aplicar cor SÓ em linhas de dados reais (DATE em col1 + status válido)
    # NÃO colorir linhas da seção 📋 (entre 📋 e a próxima 📍)
    em_secao_config = False
    for ri in range(1, ws.max_row + 1):
        v1 = ws.cell(row=ri, column=1).value
        v1s = str(v1 or "").strip()
        if ICONE_GRUPO in v1s:
            em_secao_config = False
            continue
        if ICONE_CONFIG in v1s or "Atividades configuradas" in v1s:
            em_secao_config = True
            continue
        if em_secao_config:
            continue  # nada de cor dentro da seção configurada
        if not DATE_RE.match(v1s):
            continue
        v5 = ws.cell(row=ri, column=5).value
        if v5 is None:
            continue
        bucket = _bucket(str(v5))
        if not bucket:
            continue  # status desconhecido — não colorir
        cor = _cor(bucket)
        if not cor:
            continue
        fill = PatternFill("solid", start_color=cor)
        for ci in range(1, COL_FINAL + 1):
            ws.cell(row=ri, column=ci).fill = fill

    # 7) Re-contar e atualizar texto dos cabeçalhos 📍
    counts = {}
    current_header = None
    em_secao_config = False
    for ri in range(1, ws.max_row + 1):
        v1 = ws.cell(row=ri, column=1).value
        v1s = str(v1 or "").strip()
        if ICONE_GRUPO in v1s:
            current_header = ri
            em_secao_config = False
            counts[current_header] = {
                "feitas": 0, "parciais": 0, "nao_iniciadas": 0, "perdidas": 0,
                "esperadas_faltam": 0,
                "local": re.split(r"[—\-]", v1s.replace(ICONE_GRUPO, ""), maxsplit=1)[0].strip(),
            }
            continue
        if ICONE_CONFIG in v1s or "Atividades configuradas" in v1s:
            em_secao_config = True
            continue
        if em_secao_config or current_header is None:
            continue
        if not DATE_RE.match(v1s):
            continue
        v5 = ws.cell(row=ri, column=5).value
        v5s = str(v5 or "").strip().lower()
        bucket = _bucket(str(v5) if v5 else "")
        if bucket == "ok":
            counts[current_header]["feitas"] += 1
        elif bucket == "parcial":
            counts[current_header]["parciais"] += 1
        elif bucket == "nao_feita":
            # quebra real: Perdida ≠ Não iniciada
            if v5s == "perdida":
                counts[current_header]["perdidas"] += 1
            else:
                counts[current_header]["nao_iniciadas"] += 1
        elif bucket == "esperada_falta":
            counts[current_header]["esperadas_faltam"] += 1

    for header_row, info in counts.items():
        f, p = info["feitas"], info["parciais"]
        ni, pe, ef = info["nao_iniciadas"], info["perdidas"], info["esperadas_faltam"]
        total = f + p + ni + pe
        pct = (100.0 * f / total) if total else 0.0
        partes = [f"✓ {f} Feitas", f"⚠ {p} Parciais"]
        if ni:
            partes.append(f"✗ {ni} Não iniciadas")
        if pe:
            partes.append(f"⊘ {pe} Perdidas")
        if ef:
            partes.append(f"❌ {ef} Esperadas-não-registradas")
        novo = (f"  📍  {info['local']}   —   " + "   ".join(partes)
                + f"   |   {pct:.1f}% cumprimento")
        # Info do cruzamento — se o local tem postos/*.json configurado, mostra
        # quantas atividades estavam esperadas pro dia e quantas foram feitas
        info_cruz = cruz.get(_slug(info["local"]))
        if info_cruz and info_cruz.get("esperadas_total", 0) > 0:
            esp = info_cruz["esperadas_total"]
            ok_esp = info_cruz["feitas_ok"]
            novo += f"   |   📋 {ok_esp}/{esp} avulsas esperadas feitas"
        ws.cell(row=header_row, column=1, value=novo)

    # 8) Melhorias visuais na aba Atividades — freeze + autofilter + wrap
    ws.freeze_panes = "A2"
    try:
        ws.auto_filter.ref = f"A1:{get_column_letter(COL_FINAL)}1"
    except Exception:
        pass
    # Aumentar largura da col Justificativa (col 11 = K) + wrap
    ws.column_dimensions[get_column_letter(COL_FINAL)].width = 50
    for ri in range(2, ws.max_row + 1):
        v1 = str(ws.cell(row=ri, column=1).value or "").strip()
        if DATE_RE.match(v1):
            cell = ws.cell(row=ri, column=COL_FINAL)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 9) Aba unificada "Atividades" (cruzamento limpo, cor só no status) e
    #    remoção da aba bruta linha-a-linha + da antiga "Resumo por Posto".
    #    O usuário quer UMA aba só, com menos cor.
    if "Resumo por Posto" in wb.sheetnames:
        del wb["Resumo por Posto"]
    _criar_aba_unificada(wb, dados)          # cria "Atividades — por local" (temp)
    if "Atividades" in wb.sheetnames:
        del wb["Atividades"]                 # remove a bruta linha-a-linha
    wb["Atividades — por local"].title = "Atividades"
    _mover_aba_apos_capa(wb, "Atividades")

    wb.save(xlsx_path)
    return {
        "ok": True,
        "removidas_da_rodada_anterior": n_removidas,
        "grupos_processados": len(grupos),
        "aba_unificada_criada": True,
    }


def _mover_aba_apos_capa(wb, nome):
    """Reposiciona a aba `nome` logo após 'Capa Executiva' (ou no início)."""
    if nome not in wb.sheetnames:
        return
    alvo = wb.sheetnames.index("Capa Executiva") + 1 if "Capa Executiva" in wb.sheetnames else 0
    atual = wb.sheetnames.index(nome)
    wb.move_sheet(nome, offset=alvo - atual)


# ─── Aba "Resumo por Posto" ──────────────────────────────────────────────────

COR_BG_CABLOCAL = "2E75B6"
COR_TXT_CABLOCAL = "FFFFFF"
COR_BG_HEADER = "1F4E79"

def _classificar_linhas_local(agg, cruz):
    """Constrói lista de dicts unificados (uma linha por modelo) pra renderizar.
    Cada linha tem: tipo, posto, modelo, esperado, ok, parc, nf, status, justifs."""
    rows = []
    esperadas = cruz.get("esperadas_detalhe", []) or []
    extras = cruz.get("extras", []) or []

    # Index justificativas por modelo normalizado
    just_idx = {}
    for j in agg.get("justificativas", []) or []:
        mod = _norm_modelo(j.get("modelo", ""))
        just_idx.setdefault(mod, []).append(j.get("texto", ""))

    def _justifs(modelo):
        textos = just_idx.get(_norm_modelo(modelo), [])
        # dedup mantendo ordem
        seen, uniq = set(), []
        for t in textos:
            if t and t not in seen:
                seen.add(t); uniq.append(t)
        return uniq

    # Programadas (esperadas no postos/*.json)
    for e in esperadas:
        ok = e.get("feitas_ok", 0)
        parc = e.get("parcial", 0)
        vezes = e.get("vezes", 1)
        total_dia = e.get("total_no_dia", 0)
        nf = max(0, vezes - ok - parc) if e.get("status") != "feita" else 0
        status_str = e.get("status", "nao_feita")
        if total_dia == 0 and status_str == "nao_feita":
            tag = "❌ Não registrada"
        elif status_str == "feita":
            tag = "✓ Feita"
        elif status_str == "parcial":
            tag = "⚠ Parcial"
        else:
            tag = "✗ Não feita"
        rows.append({
            "tipo": "Programada",
            "posto": e.get("posto", "—"),
            "modelo": e.get("modelo", ""),
            "esperado": vezes,
            "ok": ok, "parc": parc, "nf": nf,
            "tag": tag, "status_key": status_str,
            "justifs": _justifs(e.get("modelo", "")),
            "total_no_dia": total_dia,
        })

    # Extras (apareceram mas não estavam cadastradas)
    for ex in extras:
        ok = ex.get("ok", 0); parc = ex.get("parcial", 0); nf = ex.get("nao_feita", 0)
        if ok and not (parc or nf):
            tag = "+ Extra OK"; status_key = "feita"
        elif parc and not nf:
            tag = "+ Extra parc"; status_key = "parcial"
        else:
            tag = "+ Extra falhou"; status_key = "nao_feita"
        rows.append({
            "tipo": "Extra (não cadastrada)",
            "posto": "—",
            "modelo": ex.get("modelo", ""),
            "esperado": 0,
            "ok": ok, "parc": parc, "nf": nf,
            "tag": tag, "status_key": status_key,
            "justifs": _justifs(ex.get("modelo", "")),
            "total_no_dia": ex.get("total", ok + parc + nf),
        })

    # Sem cadastro (local sem postos/*.json mas com atividades no dia)
    if not esperadas and not extras:
        for m in agg.get("por_modelo", []) or []:
            ok = m.get("ok", 0); parc = m.get("parcial", 0); nf = m.get("nao_feita", 0)
            total = m.get("total", ok + parc + nf)
            if total == 0:
                continue
            if ok and not (parc or nf):
                tag = "✓ Feita"; status_key = "feita"
            elif parc and not nf:
                tag = "⚠ Parcial"; status_key = "parcial"
            elif ok or parc:
                tag = "⚠ Mista"; status_key = "parcial"
            else:
                tag = "✗ Não feita"; status_key = "nao_feita"
            rows.append({
                "tipo": "Sem cadastro",
                "posto": "—",
                "modelo": m.get("modelo", ""),
                "esperado": "—",
                "ok": ok, "parc": parc, "nf": nf,
                "tag": tag, "status_key": status_key,
                "justifs": _justifs(m.get("modelo", "")),
                "total_no_dia": total,
            })
    return rows


def _status_render(r):
    """Mapeia uma linha classificada → (chave_status, rótulo) pra pílula."""
    if r["tag"].startswith("❌"):
        return "nao_registrada", "não registrada"
    sk = r["status_key"]
    if sk == "feita":
        return "feita", "feita"
    if sk == "parcial":
        return "parcial", "parcial"
    return "nao_feita", "não feita"


def _criar_aba_unificada(wb, dados):
    """Cria a aba unificada (Resumo + Atividades numa só).

    Por local: atividades vindas do SISTEMA mas não cadastradas aparecem no
    topo (destaque creme, "será aprendida"); abaixo, as do CADASTRO com seu
    status. Cor só na coluna Status — o resto fica branco/zebra pra não poluir.
    """
    name = "Atividades — por local"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    headers = ["Origem", "Posto", "Modelo", "Esp.", "Feito", "Status", "Justificativa"]
    widths = [13, 18, 40, 6, 7, 16, 50]
    C_ORIGEM, C_POSTO, C_MODELO, C_ESP, C_FEITO, C_STATUS, C_JUST = range(1, 8)

    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        c.fill = PatternFill("solid", start_color=COR_BG_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    cruz_idx = dados.get("cruzamento_por_local", {})
    row = 2

    # locais ordenados por cumprimento ascendente (piores primeiro)
    locais = []
    for agg in dados.get("atividades_agg", []) or []:
        total = agg.get("total", 0)
        pct = (100.0 * agg.get("ok", 0) / total) if total else 0
        locais.append((pct, agg))
    locais.sort(key=lambda x: x[0])

    for pct, agg in locais:
        slug = agg.get("slug", "")
        nome = agg.get("nome", slug)
        linhas = _classificar_linhas_local(agg, cruz_idx.get(slug, {}))
        if not linhas:
            continue

        # sistema (extras / sem cadastro) primeiro, cadastro depois
        sistema = [r for r in linhas if r["tipo"] != "Programada"]
        cadastro = [r for r in linhas if r["tipo"] == "Programada"]
        ordenadas = sistema + cadastro

        f = sum(1 for r in cadastro if r["status_key"] == "feita")
        tot_cad = len(cadastro)
        cab = (f"📍  {nome}    {f}/{tot_cad} cadastradas feitas"
               + (f"    +{len(sistema)} do sistema" if sistema else ""))
        cell = ws.cell(row=row, column=1, value=cab)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color=COR_BG_CABLOCAL)
        cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        for idx, r in enumerate(ordenadas):
            eh_sistema = r["tipo"] != "Programada"
            modelo = r["modelo"]
            if eh_sistema:
                origem = "⬆ sistema"
                modelo = f"{modelo}  · nova, será aprendida"
            else:
                origem = "cadastro"

            ws.cell(row=row, column=C_ORIGEM, value=origem)
            ws.cell(row=row, column=C_POSTO, value=r["posto"])
            ws.cell(row=row, column=C_MODELO, value=modelo)
            ws.cell(row=row, column=C_ESP, value=r["esperado"])
            ws.cell(row=row, column=C_FEITO, value=r["ok"])
            sk, rotulo = _status_render(r)
            ws.cell(row=row, column=C_STATUS, value=rotulo)
            ws.cell(row=row, column=C_JUST,
                    value=" | ".join(r["justifs"][:2]) if r["justifs"] else "")

            # fundo da LINHA: creme leve se sistema, zebra se cadastro
            if eh_sistema:
                linha_bg = PatternFill("solid", start_color=COR_LINHA_SISTEMA)
            elif idx % 2:
                linha_bg = PatternFill("solid", start_color=COR_ZEBRA)
            else:
                linha_bg = FILL_BRANCO
            for ci in range(1, len(headers) + 1):
                ws.cell(row=row, column=ci).fill = linha_bg

            # cor SÓ na célula de status (a pílula)
            bg, fg = STATUS_CORES.get(sk, ("FFFFFF", "000000"))
            cst = ws.cell(row=row, column=C_STATUS)
            cst.fill = PatternFill("solid", start_color=bg)
            cst.font = Font(bold=True, color=fg, size=9)
            cst.alignment = Alignment(horizontal="center", vertical="center")

            # alinhamentos / fontes do resto
            ws.cell(row=row, column=C_ORIGEM).font = Font(
                color=("9A3412" if eh_sistema else "888888"), size=9,
                bold=eh_sistema)
            for ci in (C_ESP, C_FEITO):
                ws.cell(row=row, column=ci).alignment = Alignment(
                    horizontal="center", vertical="center")
            ws.cell(row=row, column=C_JUST).alignment = Alignment(
                wrap_text=True, vertical="top")
            ws.cell(row=row, column=C_POSTO).font = Font(color="888888", size=9)
            row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"


def main():
    if len(sys.argv) < 3:
        print(json.dumps({"erro": "uso: python enriquecer_atividades.py <arquivo.xlsx> <dados.json>"}))
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    json_path = Path(sys.argv[2])

    if not xlsx_path.exists():
        print(json.dumps({"erro": f"xlsx nao encontrado: {xlsx_path}"}))
        sys.exit(1)
    if not json_path.exists():
        print(json.dumps({"erro": f"json nao encontrado: {json_path}"}))
        sys.exit(1)

    try:
        dados = json.loads(json_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        print(json.dumps({"erro": f"json invalido: {e}"}))
        sys.exit(1)

    try:
        res = enriquecer(str(xlsx_path), dados)
    except PermissionError:
        print(json.dumps({"erro": f"nao consegui salvar (arquivo aberto no Excel?): {xlsx_path}"}))
        sys.exit(1)

    print(json.dumps(res, ensure_ascii=False))


if __name__ == "__main__":
    main()
