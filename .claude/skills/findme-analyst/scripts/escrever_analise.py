#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
escrever_analise.py — Escreve a aba "Análise" formatada dentro de um relatório
FindMe, a partir de um JSON com os blocos do diagnóstico.

Uso:
    python escrever_analise.py "<arquivo.xlsx>" "<analise.json>"

O JSON tem o formato {"blocos": [...]}. Os tipos de bloco estão documentados
em references/findme-domain.md (seção 6). A aba "Análise" é criada como a
primeira aba; se já existir, é substituída. O arquivo original é sobrescrito.
"""
import sys
import json
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl nao instalado. Rode: pip install openpyxl")
    sys.exit(1)

# Paleta FindMe
AZUL_ESCURO = "1F4E79"
AZUL_MEDIO = "2E75B6"
AZUL_CLARO = "D6E4F0"
BRANCO = "FFFFFF"
TEXTO = "1A1A1A"
CINZA_CLARO = "F2F2F2"

REALCE = {
    "critico": "FFC7CE",
    "atencao": "FFEB9C",
    "ok": "C6EFCE",
    "neutro": None,
    None: None,
}

NCOLS = 8  # largura padrão dos blocos de texto
THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_):
    return PatternFill("solid", start_color=hex_) if hex_ else None


def _merge(ws, row, ncols):
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)


def render(ws, blocos):
    r = 1
    max_col_usado = NCOLS

    for bloco in blocos:
        tipo = bloco.get("tipo", "")

        if tipo == "titulo":
            c = ws.cell(row=r, column=1, value=bloco.get("texto", ""))
            c.font = Font(name="Arial", bold=True, size=15, color=BRANCO)
            c.alignment = Alignment(horizontal="left", vertical="center")
            for col in range(1, NCOLS + 1):
                ws.cell(row=r, column=col).fill = _fill(AZUL_ESCURO)
            _merge(ws, r, NCOLS)
            ws.row_dimensions[r].height = 30
            r += 2

        elif tipo == "veredito":
            texto = bloco.get("texto", "")
            c = ws.cell(row=r, column=1, value=texto)
            c.font = Font(name="Arial", size=11, color=TEXTO)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for col in range(1, NCOLS + 1):
                cell = ws.cell(row=r, column=col)
                cell.fill = _fill(AZUL_CLARO)
                cell.border = BORDER
            _merge(ws, r, NCOLS)
            linhas_estim = max(2, len(texto) // (NCOLS * 11) + texto.count("\n") + 1)
            ws.row_dimensions[r].height = 16 * linhas_estim
            r += 2

        elif tipo == "secao":
            c = ws.cell(row=r, column=1, value=bloco.get("texto", ""))
            c.font = Font(name="Arial", bold=True, size=12, color=BRANCO)
            c.alignment = Alignment(horizontal="left", vertical="center")
            for col in range(1, NCOLS + 1):
                ws.cell(row=r, column=col).fill = _fill(AZUL_MEDIO)
            _merge(ws, r, NCOLS)
            ws.row_dimensions[r].height = 22
            r += 1

        elif tipo == "paragrafo":
            texto = bloco.get("texto", "")
            c = ws.cell(row=r, column=1, value=texto)
            c.font = Font(name="Arial", size=10, color=TEXTO)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            _merge(ws, r, NCOLS)
            linhas_estim = max(1, len(texto) // (NCOLS * 12) + texto.count("\n") + 1)
            ws.row_dimensions[r].height = 15 * linhas_estim
            r += 1

        elif tipo == "lista":
            for item in bloco.get("itens", []):
                c = ws.cell(row=r, column=1, value=f"•  {item}")
                c.font = Font(name="Arial", size=10, color=TEXTO)
                c.alignment = Alignment(horizontal="left", vertical="top",
                                        wrap_text=True)
                _merge(ws, r, NCOLS)
                linhas_estim = max(1, len(str(item)) // (NCOLS * 12) + 1)
                ws.row_dimensions[r].height = 15 * linhas_estim
                r += 1
            r += 1

        elif tipo == "tabela":
            colunas = bloco.get("colunas", [])
            linhas = bloco.get("linhas", [])
            realces = bloco.get("realces", [])
            ncols = max(len(colunas), 1)
            max_col_usado = max(max_col_usado, ncols)
            # cabeçalho
            for ci, nome in enumerate(colunas, start=1):
                cell = ws.cell(row=r, column=ci, value=nome)
                cell.font = Font(name="Arial", bold=True, size=10, color=BRANCO)
                cell.fill = _fill(AZUL_ESCURO)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border = BORDER
            ws.row_dimensions[r].height = 20
            r += 1
            # dados
            for li, linha in enumerate(linhas):
                realce = realces[li] if li < len(realces) else None
                fill = _fill(REALCE.get(realce))
                for ci in range(ncols):
                    val = linha[ci] if ci < len(linha) else ""
                    cell = ws.cell(row=r, column=ci + 1, value=val)
                    cell.font = Font(name="Arial", size=10, color=TEXTO)
                    cell.alignment = Alignment(
                        horizontal="left" if ci == 0 else "center",
                        vertical="center", wrap_text=True)
                    cell.border = BORDER
                    if fill:
                        cell.fill = fill
                    elif li % 2 == 1:
                        cell.fill = _fill(CINZA_CLARO)
                r += 1
            r += 1

        else:
            # bloco desconhecido — ignora silenciosamente para não travar
            continue

    # larguras de coluna
    ws.column_dimensions["A"].width = 34
    for col in range(2, max_col_usado + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.sheet_view.showGridLines = False


def main():
    if len(sys.argv) < 3:
        print("uso: python escrever_analise.py <arquivo.xlsx> <analise.json>")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    json_path = Path(sys.argv[2])

    if not xlsx_path.exists():
        print(f"ERRO: arquivo nao encontrado: {xlsx_path}")
        sys.exit(1)
    if not json_path.exists():
        print(f"ERRO: json nao encontrado: {json_path}")
        sys.exit(1)

    try:
        analise = json.loads(json_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        print(f"ERRO: json invalido: {e}")
        sys.exit(1)

    blocos = analise.get("blocos", [])
    if not blocos:
        print("ERRO: o json nao tem a chave 'blocos' ou ela esta vazia.")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        print(f"ERRO: falha ao abrir o xlsx: {e}")
        sys.exit(1)

    if "Análise" in wb.sheetnames:
        del wb["Análise"]
    ws = wb.create_sheet("Análise", 0)
    ws.sheet_properties.tabColor = AZUL_ESCURO

    render(ws, blocos)

    try:
        wb.save(xlsx_path)
    except PermissionError:
        print(f"ERRO: nao foi possivel salvar — o arquivo esta aberto no Excel? "
              f"Feche-o e rode de novo: {xlsx_path}")
        sys.exit(1)

    print(f"OK: aba 'Análise' escrita em {xlsx_path} ({len(blocos)} blocos)")


if __name__ == "__main__":
    main()
