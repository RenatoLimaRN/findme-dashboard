#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ler_relatorio.py — Detecta o tipo de relatório FindMe e extrai um JSON
normalizado, pronto para o diagnóstico operacional.

Uso básico:
    python ler_relatorio.py "<arquivo.xlsx>"

Modo D-1 (uso principal do skill):
    python ler_relatorio.py "<arquivo.xlsx>" \\
        --data 2026-05-15 \\
        --postos-dir "<raiz do projeto>/postos" \\
        --historico-dir "<skill>/historico"

Quando --data é dado: filtra as atividades para apenas aquele dia.
Quando --postos-dir é dado: carrega o registro local de avulsas esperadas
e cruza esperado × feito por local (precisa de --data).
Quando --historico-dir é dado: anexa snapshots recentes por local.

Em caso de erro: imprime {"erro": "..."} e sai com código 1.
"""
import sys
import json
import re
import glob
import argparse
import unicodedata
from datetime import date
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    print(json.dumps({"erro": "openpyxl nao instalado. Rode: pip install openpyxl"}))
    sys.exit(1)

DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
DIAS_PT_ORDEM = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]  # Mon..Sun

# Status (rótulo no relatório) -> balde operacional
BALDE = {
    "completa": "ok",
    "incompleta": "parcial",
    "incompleta c/ justif.": "parcial",
    "incompleta c/ justificativa": "parcial",
    "perdida": "nao_feita",
    "nao iniciada": "nao_feita",
    "não iniciada": "nao_feita",
}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _pct(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace("%", "").replace(",", "."))
    except ValueError:
        return None


def _int(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(float(str(v).strip().replace(",", ".")))
    except ValueError:
        return 0


def _txt(v):
    return "" if v is None else str(v).strip()


def _rows(ws):
    return [[c.value for c in row] for row in ws.iter_rows()]


def _strip_accents(s):
    s = unicodedata.normalize("NFD", s or "")
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def _slug(nome):
    """Slug consistente com o que findme_programacao.py usa para nomear postos."""
    s = _strip_accents(nome)
    s = re.sub(r"^\s*CONDOMINIO\s+", "", s, flags=re.I).strip()
    s = re.sub(r"[^A-Za-z0-9]+", "_", s.lower()).strip("_")
    return s


def _norm_modelo(m):
    return re.sub(r"\s+", " ", _strip_accents(m).upper().strip())


def _norm_dia(d):
    s = _strip_accents(d).upper().strip()[:3]
    # Sáb vira "SAB", Dom vira "DOM", etc.
    return s


def _weekday_pt(yyyy_mm_dd):
    try:
        d = date.fromisoformat(yyyy_mm_dd)
        return DIAS_PT_ORDEM[d.weekday()]
    except Exception:
        return None


def _data_br(yyyy_mm_dd):
    try:
        return date.fromisoformat(yyyy_mm_dd).strftime("%d/%m/%Y")
    except Exception:
        return None


# ─── Registro de avulsas esperadas (postos/*.json) ────────────────────────────

def _carregar_postos_registro(postos_dir):
    """Carrega postos/*.json. Templates vazios (sem atividades configuradas)
    são ignorados — assim o fuzzy match consegue casar a variação curta do
    nome do local com o registro real que tem as atividades preenchidas."""
    out = {}
    vazios = []
    if not postos_dir:
        return out, vazios
    for fp in sorted(glob.glob(str(Path(postos_dir) / "*.json"))):
        try:
            with open(fp, encoding="utf-8") as f:
                data = json.load(f)
            local = data.get("local")
            if not local:
                continue
            total_ativ = sum(len(p.get("atividades", []))
                             for p in data.get("postos", []))
            if total_ativ == 0:
                vazios.append({"path": fp, "local": local, "slug": _slug(local)})
                continue
            out[_slug(local)] = {"path": fp, "local": local, "registro": data}
        except Exception:
            pass
    return out, vazios


def _longest_common_prefix(a, b):
    n = min(len(a), len(b))
    for i in range(n):
        if a[i] != b[i]:
            return i
    return n


def _match_registro(group_slug, registros_by_slug, min_overlap=6):
    """Casa o slug do grupo (vindo do relatório) com o slug do registro
    (vindo do postos/*.json). Exato primeiro, depois prefixo comum mais longo
    com pelo menos `min_overlap` chars.

    Por que: o mesmo local aparece com nomes diferentes em partes diferentes
    do relatório FindMe (ex.: 'CONDOMÍNIO VIVAZ PRIME - RIO BONITO' na Capa
    e só 'CONDOMÍNIO VIVAZ PRIME' na aba Atividades). Os dois devem casar
    com o mesmo registro postos/condominio_vivaz_prime_rio_bonito.json.

    Retorna (slug_casado, info) ou (None, None) se não casar.
    """
    if not group_slug:
        return (None, None)
    if group_slug in registros_by_slug:
        return (group_slug, registros_by_slug[group_slug])
    best_slug = None
    best_overlap = 0
    for r_slug in registros_by_slug:
        overlap = _longest_common_prefix(group_slug, r_slug)
        # exige que o prefixo bata "até o fim" de pelo menos um dos lados —
        # senão "vivaz_prime" e "vivaz_norte" casariam só pelo "vivaz_"
        if overlap >= min_overlap and (overlap == len(group_slug) or overlap == len(r_slug)):
            if overlap > best_overlap:
                best_overlap = overlap
                best_slug = r_slug
    if best_slug:
        return (best_slug, registros_by_slug[best_slug])
    return (None, None)


def _calcular_cruzamento(local_agg, registro, weekday):
    """Esperado × feito do dia. local_agg pode ser None (sem atividades no dia)."""
    esperadas = []
    wkd_norm = _norm_dia(weekday) if weekday else ""
    for posto in registro.get("postos", []):
        for ativ in posto.get("atividades", []):
            dias_norm = [_norm_dia(d) for d in ativ.get("dias", [])]
            if wkd_norm and wkd_norm in dias_norm:
                esperadas.append({
                    "modelo": ativ.get("modelo", ""),
                    "vezes": int(ativ.get("vezes", 1) or 1),
                    "posto": posto.get("posto", ""),
                })

    feito_idx = {}
    if local_agg:
        for pm in local_agg.get("por_modelo", []):
            feito_idx[_norm_modelo(pm["modelo"])] = pm

    esperadas_detalhe = []
    feitas_ok = parciais = perdidas = esperadas_total = 0
    modelos_esperados_norm = set()

    for e in esperadas:
        nm = _norm_modelo(e["modelo"])
        modelos_esperados_norm.add(nm)
        vezes = e["vezes"]
        pm = feito_idx.get(nm)
        ok = pm["ok"] if pm else 0
        parcial = pm["parcial"] if pm else 0
        total = pm["total"] if pm else 0
        esperadas_total += vezes
        if ok >= vezes:
            status = "feita"
            feitas_ok += vezes
        elif ok + parcial >= vezes:
            status = "parcial"
            feitas_ok += ok
            parciais += vezes - ok
        else:
            status = "nao_feita"
            feitas_ok += ok
            parciais += parcial
            perdidas += max(0, vezes - ok - parcial)
        esperadas_detalhe.append({
            "modelo": e["modelo"], "posto": e["posto"], "vezes": vezes,
            "feitas_ok": ok, "parcial": parcial, "total_no_dia": total,
            "status": status,
        })

    extras = []
    if local_agg:
        for pm in local_agg.get("por_modelo", []):
            if _norm_modelo(pm["modelo"]) not in modelos_esperados_norm:
                extras.append({
                    "modelo": pm["modelo"], "total": pm["total"],
                    "ok": pm["ok"], "parcial": pm["parcial"],
                    "nao_feita": pm["nao_feita"],
                })

    return {
        "esperadas_total": esperadas_total,
        "feitas_ok": feitas_ok,
        "parciais": parciais,
        "perdidas": perdidas,
        "esperadas_detalhe": esperadas_detalhe,
        "extras": extras,
    }


# ─── Histórico ────────────────────────────────────────────────────────────────

def _carregar_historico(hist_dir, slug_local, days):
    out = []
    if not hist_dir:
        return out
    base = Path(hist_dir)
    if not base.exists():
        return out
    for ddir in sorted(base.iterdir(), reverse=True):
        if not ddir.is_dir() or not re.match(r"^\d{4}-\d{2}-\d{2}$", ddir.name):
            continue
        f = ddir / f"{slug_local}.json"
        if f.exists():
            try:
                out.append(json.loads(f.read_text(encoding="utf-8")))
            except Exception:
                pass
        if len(out) >= days:
            break
    return out


# ─── Relatório GERAL (findme_programacao.py) ──────────────────────────────────

def parse_geral(wb, data_alvo=None):
    out = {"tipo": "geral_programacao", "periodo": None,
           "kpis_gerais": {}, "locais": [], "atividades_agg": []}
    data_alvo_br = _data_br(data_alvo) if data_alvo else None

    if "Capa Executiva" in wb.sheetnames:
        rows = _rows(wb["Capa Executiva"])
        for r in rows[:4]:
            for cell in r:
                t = _txt(cell)
                if "Período" in t or "Periodo" in t:
                    out["periodo"] = t
        # KPIs: rótulo na linha L, valor na L-1, casados por coluna (cells mescladas)
        for i, r in enumerate(rows):
            idx = {}
            for ci, c in enumerate(r):
                t = _txt(c)
                if "Eficiência Geral" in t or "Eficiencia Geral" in t:
                    idx["pct"] = ci
                elif "Locais Monitorados" in t:
                    idx["mon"] = ci
                elif "Críticos" in t or "Criticos" in t:
                    idx["crit"] = ci
                elif "Não Feitas" in t or "Nao Feitas" in t:
                    idx["naofeitas"] = ci
            if "pct" in idx and i > 0:
                vals = rows[i - 1]

                def _g(key, fn):
                    ci = idx.get(key)
                    return fn(vals[ci]) if ci is not None and ci < len(vals) else None

                out["kpis_gerais"] = {
                    "pct_eficiencia_geral": _g("pct", _pct),
                    "locais_monitorados": _g("mon", _int),
                    "locais_criticos": _g("crit", _int),
                    "atividades_nao_feitas": _g("naofeitas", _int),
                }
                break
        vistos = {}
        for i, r in enumerate(rows):
            if any(_txt(c) == "Local / Posto" for c in r):
                for r2 in rows[i + 1:]:
                    if not r2 or r2[0] is None or not isinstance(r2[0], (int, float)):
                        break
                    nome = _txt(r2[1])
                    if not nome:
                        break
                    vistos[nome] = {
                        "nome": nome, "slug": _slug(nome),
                        "ok": _int(r2[2]), "parcial": _int(r2[3]),
                        "nao_feita": _int(r2[4]), "total": _int(r2[5]),
                        "pct_cumprimento": _pct(r2[6]),
                    }
        out["locais"] = list(vistos.values())

    if "Atividades" in wb.sheetnames:
        rows = _rows(wb["Atividades"])
        agg = {}
        local_atual = "(sem local)"
        for r in rows:
            if not r:
                continue
            c0 = _txt(r[0])
            if "📍" in c0:
                m = re.split(r"[—\-]", c0.replace("📍", ""), maxsplit=1)
                local_atual = m[0].strip() if m else c0.strip()
                continue
            if not DATE_RE.match(c0):
                continue
            if data_alvo_br and c0 != data_alvo_br:
                continue  # filtra para o dia-alvo
            turno = _txt(r[2]) if len(r) > 2 else ""
            modelo_raw = _txt(r[3]) if len(r) > 3 else ""
            status = _txt(r[4]).lower() if len(r) > 4 else ""
            balde = BALDE.get(status, "nao_feita")
            is_avulsa = modelo_raw.upper().startswith("[AVULSA]")
            modelo = re.sub(r"^\[AVULSA\]\s*", "", modelo_raw).strip() or "(sem modelo)"

            d = agg.setdefault(local_atual, {
                "nome": local_atual, "slug": _slug(local_atual),
                "ok": 0, "parcial": 0, "nao_feita": 0,
                "total": 0, "avulsas": 0, "por_modelo": {}, "por_turno": {}})
            d[balde] += 1
            d["total"] += 1
            if is_avulsa:
                d["avulsas"] += 1
            pm = d["por_modelo"].setdefault(
                modelo, {"modelo": modelo, "avulsa": is_avulsa,
                         "ok": 0, "parcial": 0, "nao_feita": 0, "total": 0})
            pm[balde] += 1
            pm["total"] += 1
            pt = d["por_turno"].setdefault(
                turno or "(sem turno)",
                {"turno": turno or "(sem turno)",
                 "ok": 0, "parcial": 0, "nao_feita": 0, "total": 0})
            pt[balde] += 1
            pt["total"] += 1

        for d in agg.values():
            d["por_modelo"] = sorted(d["por_modelo"].values(),
                                     key=lambda x: -x["nao_feita"])
            d["por_turno"] = sorted(d["por_turno"].values(),
                                    key=lambda x: -x["nao_feita"])
        out["atividades_agg"] = list(agg.values())

    return out


# ─── Relatório de KPIs (findme_dashboard.py) ──────────────────────────────────

def parse_dashboard(wb):
    out = {"tipo": "dashboard_kpi", "periodo": None, "locais": [],
           "atividades_resumo": {}, "atividades_por_mes": [],
           "checklists": {"kpis": {}, "itens": []},
           "justificativas": [], "avulsas_perdidas": []}

    if "Locais" in wb.sheetnames:
        for r in _rows(wb["Locais"]):
            if r and _txt(r[0]) and _txt(r[0]) not in ("UUID",) and "Locais" not in _txt(r[0]):
                if len(r) >= 5 and _txt(r[1]):
                    nome = _txt(r[1])
                    out["locais"].append({
                        "uuid": _txt(r[0]), "nome": nome, "slug": _slug(nome),
                        "cliente": _txt(r[2]), "regiao": _txt(r[3]),
                        "status": _txt(r[4])})

    if "Atividades Resumo" in wb.sheetnames:
        for r in _rows(wb["Atividades Resumo"]):
            if r and _txt(r[0]) and len(r) > 1 and r[1] is not None:
                lbl = _txt(r[0])
                if lbl and lbl not in ("Indicador",) and "Resumo" not in lbl:
                    out["atividades_resumo"][lbl] = r[1]

    if "Atividades por Mês" in wb.sheetnames:
        for r in _rows(wb["Atividades por Mês"]):
            if r and _txt(r[0]) and _txt(r[0]) not in ("Mês",) \
               and "Evolução" not in _txt(r[0]) and "TOTAL" not in _txt(r[0]).upper():
                out["atividades_por_mes"].append({
                    "mes": _txt(r[0]),
                    "total": _int(r[1]) if len(r) > 1 else 0,
                    "checkins_esperados": _int(r[2]) if len(r) > 2 else 0,
                    "checkins_feitos": _int(r[3]) if len(r) > 3 else 0,
                    "eficiencia": _pct(r[4]) if len(r) > 4 else None})

    if "Checklists" in wb.sheetnames:
        for r in _rows(wb["Checklists"]):
            if not r or not _txt(r[0]):
                continue
            lbl = _txt(r[0])
            if len(r) >= 4 and _txt(r[3]) and lbl not in ("Nome do Item",):
                out["checklists"]["itens"].append({
                    "nome": lbl, "checklist": _txt(r[1]),
                    "ocorrencias": _int(r[2]), "tipo": _txt(r[3])})
            elif len(r) > 1 and isinstance(r[1], (int, float)) \
                    and "Resumo" not in lbl and lbl not in ("Indicador",):
                out["checklists"]["kpis"][lbl] = r[1]

    if "Justificativas" in wb.sheetnames:
        for r in _rows(wb["Justificativas"]):
            if r and _txt(r[0]) and len(r) > 1 and r[1] is not None \
               and _txt(r[0]) not in ("Justificativa",) \
               and "TOTAL" not in _txt(r[0]).upper():
                out["justificativas"].append({
                    "justificativa": _txt(r[0]),
                    "total": _int(r[1]) if len(r) > 1 else 0,
                    "incompletas": _int(r[2]) if len(r) > 2 else 0,
                    "perdidas": _int(r[3]) if len(r) > 3 else 0})

    if "Avulsas Perdidas" in wb.sheetnames:
        for r in _rows(wb["Avulsas Perdidas"]):
            if r and _txt(r[0]) and _txt(r[0]) not in ("ID",) \
               and "Avulsas" not in _txt(r[0]) and "Nenhuma" not in _txt(r[0]):
                out["avulsas_perdidas"].append({
                    "id": _txt(r[0]), "data": _txt(r[1]) if len(r) > 1 else "",
                    "posto": _txt(r[4]) if len(r) > 4 else "",
                    "modelo": _txt(r[5]) if len(r) > 5 else "",
                    "local": _txt(r[6]) if len(r) > 6 else ""})

    return out


# ─── Enriquecimento (cruzamento + histórico + postos sem registro) ────────────

def enriquecer(data, postos_dir, historico_dir, historico_dias, data_alvo):
    """Adiciona cruzamento_por_local, historico_por_local, postos_sem_registro."""
    data["dia_alvo"] = data_alvo
    data["dia_semana"] = _weekday_pt(data_alvo) if data_alvo else None
    data["cruzamento_por_local"] = {}
    data["historico_por_local"] = {}
    data["postos_sem_registro"] = []

    registros, templates_vazios = _carregar_postos_registro(postos_dir)
    data["registros_carregados"] = sorted(registros.keys())
    data["postos_template_vazio"] = templates_vazios  # precisam ser preenchidos

    # União de locais para enriquecer: locais (Capa) + atividades_agg + dashboard locais
    locais_unicos = {}
    for src in ("locais", "atividades_agg"):
        for loc in data.get(src, []) or []:
            slug = loc.get("slug")
            if slug and slug not in locais_unicos:
                locais_unicos[slug] = loc.get("nome", "")

    data["matches_fuzzy"] = []  # registros casados via fuzzy (pra transparência)

    for slug, nome in locais_unicos.items():
        # Histórico
        if historico_dir:
            hist = _carregar_historico(historico_dir, slug, historico_dias)
            if hist:
                data["historico_por_local"][slug] = hist

        # Cruzamento (precisa de --data e do registro do local) — com fuzzy match
        matched_slug, reg_info = _match_registro(slug, registros)
        if reg_info is None:
            data["postos_sem_registro"].append({"nome": nome, "slug": slug})
            continue

        if matched_slug != slug:
            data["matches_fuzzy"].append({
                "grupo_slug": slug,
                "grupo_nome": nome,
                "registro_slug": matched_slug,
                "registro_local": reg_info["local"],
            })

        if data_alvo:
            local_agg = next(
                (a for a in data.get("atividades_agg", []) if a.get("slug") == slug),
                None,
            )
            cruz = _calcular_cruzamento(
                local_agg, reg_info["registro"], data["dia_semana"])
            cruz["registro_path"] = reg_info["path"]
            cruz["registro_local"] = reg_info["local"]
            data["cruzamento_por_local"][slug] = cruz

    return data


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description="Extrai um JSON normalizado de um relatório FindMe.")
    p.add_argument("xlsx", help="Caminho do .xlsx do relatório")
    p.add_argument("--data", help="Dia-alvo YYYY-MM-DD (D-1 normalmente). Filtra atividades para esse dia.")
    p.add_argument("--postos-dir", dest="postos_dir",
                   help="Pasta com os postos/*.json (registro local de avulsas esperadas).")
    p.add_argument("--historico-dir", dest="historico_dir",
                   help="Pasta com o histórico de snapshots (historico/YYYY-MM-DD/<slug>.json).")
    p.add_argument("--historico-dias", dest="historico_dias", type=int, default=14,
                   help="Quantos dias de histórico recente carregar (default: 14).")
    args = p.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print(json.dumps({"erro": f"arquivo nao encontrado: {path}"}))
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(json.dumps({"erro": f"falha ao abrir o arquivo: {e}"}))
        sys.exit(1)

    nomes = set(wb.sheetnames)
    if {"Capa Executiva"} & nomes or any("Ranking" in n for n in nomes):
        data = parse_geral(wb, data_alvo=args.data)
    elif {"Atividades Resumo", "Justificativas"} & nomes:
        data = parse_dashboard(wb)
    else:
        print(json.dumps({"erro": "tipo de relatorio nao reconhecido",
                          "abas_encontradas": list(nomes)}))
        sys.exit(1)

    data["arquivo"] = str(path)
    enriquecer(data, args.postos_dir, args.historico_dir,
               args.historico_dias, args.data)

    print(json.dumps(data, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
