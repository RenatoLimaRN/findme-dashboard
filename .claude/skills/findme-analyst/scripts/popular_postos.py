#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
popular_postos.py — Sugere/aplica cadastro de atividades pros postos vazios
em postos/*.json, baseado no histórico de relatórios xlsx.

LÓGICA:
  1. Varre TODOS os relatorios/<periodo>/GERAL_*.xlsx do projeto
  2. Pra cada local, agrega quais modelos apareceram em quais dias da semana
     (independente do status — uma "Não iniciada" também conta como cadastrada
     no FindMe)
  3. Pra cada posto VAZIO em postos/<slug>.json:
     - Propõe atividades cujos modelos apareceram em >= threshold dias distintos
     - Categoriza por nome: "Ronda" / "Limpeza" / "Alertas" / "Outros"
     - dias = dias da semana em que apareceu pelo menos 1 vez
     - vezes = média (round) de aparições por dia da semana
     - marca cada atividade com "_sugerido": true + "_baseado_em": "X dias"
  4. Postos JÁ cadastrados (atividades não-vazia) NÃO são sobrescritos.

USO:
    python popular_postos.py <relatorios-dir> <postos-dir>          # preview
    python popular_postos.py <relatorios-dir> <postos-dir> --aplicar
    [--threshold-dias 2]   minimo de dias distintos pra sugerir (default 2)
    [--apenas-vazios]      só preenche templates vazios (default true)

Exemplo:
    python .claude/skills/findme-analyst/scripts/popular_postos.py \
        relatorios postos --aplicar
"""
import argparse
import glob
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl nao instalado")
    sys.exit(1)

DATE_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{4})$")
DIAS_PT_ORDEM = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]


def _slug(nome):
    s = unicodedata.normalize("NFD", nome or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"^\s*CONDOMINIO\s+", "", s, flags=re.I).strip()
    return re.sub(r"[^A-Za-z0-9]+", "_", s.lower()).strip("_")


def _weekday_pt(dd_mm_yyyy):
    m = DATE_RE.match(dd_mm_yyyy)
    if not m:
        return None
    try:
        d = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        return DIAS_PT_ORDEM[d.weekday()]
    except Exception:
        return None


def _categorizar_posto(modelo: str) -> str:
    """Heurística simples: pelo nome, decide qual 'posto' a atividade pertence."""
    m = modelo.upper()
    if "RONDA" in m or "PATRULHA" in m or "PERÍMETRO" in m or "SETORIAL" in m:
        return "Ronda"
    if "LIMPEZA" in m or "LIMPAR" in m or "HALL" in m or "PISCINA" in m \
            or "ÁREA" in m and ("EXTERN" in m or "COMUM" in m):
        return "Limpeza"
    if "ALERTA" in m or "ALERT" in m:
        return "Alertas"
    if "VISTORIA" in m:
        return "Vistorias"
    return "Outros"


def _op_tipo_por_posto(posto: str) -> str:
    """Mapeia o nome do posto pro op_tipo padrão do FindMe."""
    return {
        "Ronda": "Vigilante",
        "Limpeza": "Limpeza",
        "Alertas": "Vigilante",
        "Vistorias": "Vigilante",
    }.get(posto, "N/C")


# ─── Varredura dos relatórios ────────────────────────────────────────────────

def varrer_relatorios(relatorios_dir: Path):
    """Agrega: {slug_local: {nome_real, modelos: {modelo: {dia_semana: contagem, dias_distintos: set}}}}."""
    agg = defaultdict(lambda: {"nome": "", "modelos": defaultdict(
        lambda: {"por_dia": defaultdict(int), "dias_unicos": set(),
                 "datas_unicas": set()})})
    arquivos = sorted(glob.glob(str(relatorios_dir / "**" / "GERAL_*.xlsx"),
                                recursive=True))
    print(f"  varrendo {len(arquivos)} relatório(s)...", file=sys.stderr)
    n_atividades = 0
    for fp in arquivos:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
        except Exception as e:
            print(f"  ⚠ erro abrindo {fp}: {e}", file=sys.stderr)
            continue
        if "Atividades" not in wb.sheetnames:
            continue
        ws = wb["Atividades"]
        local_atual = None
        for ri in range(1, ws.max_row + 1):
            v = ws.cell(row=ri, column=1).value
            s = str(v or "").strip()
            if "📍" in s:
                # cabeçalho de grupo → extrai nome do local
                m = re.split(r"[—\-]", s.replace("📍", ""), maxsplit=1)
                local_atual = m[0].strip() if m else s.strip()
                continue
            if not local_atual:
                continue
            if not DATE_RE.match(s):
                continue
            data = s
            wd = _weekday_pt(data)
            if not wd:
                continue
            modelo_raw = str(ws.cell(row=ri, column=4).value or "").strip()
            if not modelo_raw:
                continue
            # pula avulsas — são pontuais, não cadastradas
            if modelo_raw.upper().startswith("[AVULSA]"):
                continue
            # pula esperadas-nao-registradas que eu mesmo inseri
            if "[ESPERADA — NÃO REGISTRADA]" in modelo_raw or \
                    "[ESPERADA - NAO REGISTRADA]" in modelo_raw.upper():
                continue
            slug = _slug(local_atual)
            entry = agg[slug]
            if not entry["nome"]:
                entry["nome"] = local_atual
            mod = entry["modelos"][modelo_raw]
            mod["por_dia"][wd] += 1
            mod["dias_unicos"].add(wd)
            mod["datas_unicas"].add(data)
            n_atividades += 1
    print(f"  ✓ {n_atividades} atividades agregadas em {len(agg)} locais",
          file=sys.stderr)
    return agg


# ─── Geração de propostas ────────────────────────────────────────────────────

def gerar_proposta(modelos_agg: dict, threshold_dias: int = 2) -> list:
    """Retorna lista de postos com atividades sugeridas."""
    # 1) Pra cada modelo, decide se entra (presente em >= threshold dias distintos)
    qualificados = []
    for modelo_raw, info in modelos_agg.items():
        if len(info["datas_unicas"]) < threshold_dias:
            continue
        # dias = lista de dias da semana ordenada por DIAS_PT_ORDEM
        dias = sorted(info["dias_unicos"], key=lambda d: DIAS_PT_ORDEM.index(d))
        # vezes = média (rounded) por dia em que aparece
        total = sum(info["por_dia"].values())
        n_dias_distintos = len(info["datas_unicas"])
        # vezes por dia da semana (média)
        vezes_media = max(1, round(total / max(n_dias_distintos, 1)))
        qualificados.append({
            "modelo": modelo_raw,
            "dias": dias,
            "vezes": vezes_media,
            "_sugerido": True,
            "_baseado_em": f"{n_dias_distintos} dia(s) de histórico, {total} aparições",
            "_posto_inferido": _categorizar_posto(modelo_raw),
        })

    # 2) Agrupa por posto inferido
    por_posto = defaultdict(list)
    for q in qualificados:
        por_posto[q.pop("_posto_inferido")].append(q)

    # 3) Monta lista de postos
    postos = []
    for nome_posto, ativs in por_posto.items():
        postos.append({
            "posto": nome_posto,
            "op_tipo": _op_tipo_por_posto(nome_posto),
            "atividades": sorted(ativs, key=lambda a: a["modelo"]),
        })
    return postos


# ─── Aplicação nos arquivos postos/ ──────────────────────────────────────────

def carregar_postos_json(postos_dir: Path):
    """Retorna dict {slug: {path, data, esta_vazio}}."""
    out = {}
    for fp in sorted(postos_dir.glob("*.json")):
        try:
            data = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            continue
        slug = _slug(data.get("local", ""))
        if not slug:
            slug = fp.stem
        total_ativ = sum(len(p.get("atividades", []) or [])
                         for p in (data.get("postos") or []))
        out[slug] = {"path": fp, "data": data, "esta_vazio": total_ativ == 0}
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("relatorios_dir")
    ap.add_argument("postos_dir")
    ap.add_argument("--aplicar", action="store_true",
                    help="Escreve nos arquivos. Sem isso, só imprime preview.")
    ap.add_argument("--threshold-dias", type=int, default=2,
                    help="Modelo precisa aparecer em N+ dias distintos pra entrar (default 2)")
    ap.add_argument("--apenas-vazios", action="store_true", default=True)
    args = ap.parse_args()

    rel_dir = Path(args.relatorios_dir)
    postos_dir = Path(args.postos_dir)

    if not rel_dir.exists():
        print(f"ERRO: {rel_dir} nao existe", file=sys.stderr)
        sys.exit(1)
    if not postos_dir.exists():
        print(f"ERRO: {postos_dir} nao existe", file=sys.stderr)
        sys.exit(1)

    print("=== 1) varrendo relatórios ===", file=sys.stderr)
    agg = varrer_relatorios(rel_dir)

    print("\n=== 2) carregando postos/ ===", file=sys.stderr)
    postos_atuais = carregar_postos_json(postos_dir)
    n_vazios = sum(1 for v in postos_atuais.values() if v["esta_vazio"])
    print(f"  {len(postos_atuais)} postos/*.json (vazios: {n_vazios})",
          file=sys.stderr)

    print("\n=== 3) gerando propostas ===", file=sys.stderr)
    resumo = []
    n_aplicados = 0
    n_pulados_sem_template = 0
    n_pulados_ja_preenchido = 0
    for slug, info in sorted(agg.items()):
        modelos = info["modelos"]
        nome_local = info["nome"]
        proposta_postos = gerar_proposta(modelos, args.threshold_dias)
        total_ativ_propostas = sum(len(p["atividades"]) for p in proposta_postos)

        # Template precisa existir
        if slug not in postos_atuais:
            resumo.append({
                "slug": slug, "nome": nome_local,
                "acao": "PULADO (sem template)",
                "propostas": total_ativ_propostas,
            })
            n_pulados_sem_template += 1
            continue

        atual = postos_atuais[slug]
        if not atual["esta_vazio"] and args.apenas_vazios:
            resumo.append({
                "slug": slug, "nome": nome_local,
                "acao": "PULADO (já cadastrado)",
                "propostas": total_ativ_propostas,
            })
            n_pulados_ja_preenchido += 1
            continue

        if total_ativ_propostas == 0:
            resumo.append({
                "slug": slug, "nome": nome_local,
                "acao": "SEM PROPOSTA (poucos dados)",
                "propostas": 0,
            })
            continue

        # Aplica (ou só preview)
        novo_data = dict(atual["data"])
        novo_data["postos"] = proposta_postos
        novo_data.pop("_instrucao", None)  # template tinha instrução; remove
        if args.aplicar:
            atual["path"].write_text(
                json.dumps(novo_data, ensure_ascii=False, indent=2),
                encoding="utf-8")
            n_aplicados += 1
        resumo.append({
            "slug": slug, "nome": nome_local,
            "acao": "APLICADO" if args.aplicar else "PREVIEW",
            "propostas": total_ativ_propostas,
            "postos": proposta_postos,
        })

    # Imprime resumo
    print("\n=== 4) RESUMO ===", file=sys.stderr)
    for r in resumo:
        print(f"  {r['acao']:30s}  {r['nome']:38s}  ({r['propostas']} atividade(s))",
              file=sys.stderr)

    out = {
        "modo": "aplicar" if args.aplicar else "preview",
        "threshold_dias": args.threshold_dias,
        "totais": {
            "locais_no_historico": len(agg),
            "postos_template_existentes": len(postos_atuais),
            "aplicados": n_aplicados,
            "pulados_sem_template": n_pulados_sem_template,
            "pulados_ja_preenchido": n_pulados_ja_preenchido,
        },
        "detalhes": resumo,
    }
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
