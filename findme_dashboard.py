#!/usr/bin/env python3
"""
FindMe Dashboard — Exportador de Indicadores
Consome a Dashboard Service API v2.0 e gera relatório Excel.

Uso:
    python findme_dashboard.py

Pré-requisito:
    pip install -r requirements.txt
    Editar config.json com email, password e UUIDs dos locais.
"""

import json
import sys
import requests
from datetime import date, timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constantes ───────────────────────────────────────────────────────────────

BASE_DASHBOARD = "https://dashboard-production.findme.id"
BASE_AUTH      = "https://production.api.findme.id"

# Paleta de cores
C_DARK_BLUE  = "1F4E79"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "D6E4F0"
C_WHITE      = "FFFFFF"
C_TEXT_BLUE  = "1F4E79"

H_FILL  = PatternFill("solid", start_color=C_DARK_BLUE)
T_FILL  = PatternFill("solid", start_color=C_MID_BLUE)
A_FILL  = PatternFill("solid", start_color=C_LIGHT_BLUE)
H_FONT  = Font(name="Arial", bold=True, color=C_WHITE, size=10)
T_FONT  = Font(name="Arial", bold=True, color=C_WHITE, size=10)
D_FONT  = Font(name="Arial", size=10)
TT_FONT = Font(name="Arial", bold=True, size=12, color=C_TEXT_BLUE)

THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ─── Autenticação e chamadas à API ────────────────────────────────────────────

def login(email: str, password: str) -> str:
    r = requests.post(
        f"{BASE_AUTH}/v3/settings/login",
        json={"email": email, "password": password},
        timeout=30,
    )
    r.raise_for_status()
    data = r.json()
    token = (
        data.get("token")
        or data.get("access_token")
        or (data.get("data") or {}).get("token")
        or (data.get("data") or {}).get("access_token")
    )
    if not token:
        raise ValueError(f"Token não encontrado na resposta de login: {data}")
    return token


def api_get(token: str, path: str):
    r = requests.get(
        f"{BASE_DASHBOARD}{path}",
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()


def api_post(token: str, path: str, body: dict, params: dict = None):
    try:
        r = requests.post(
            f"{BASE_DASHBOARD}{path}",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
            json=body,
            params=params,
            timeout=30,
        )
        r.raise_for_status()
        result = r.json()
        return result if result is not None else {}
    except requests.HTTPError as e:
        print(f"    ⚠  HTTP {e.response.status_code} em POST {path}")
        return {}
    except Exception as e:
        print(f"    ⚠  Erro em POST {path}: {e}")
        return {}


# ─── Configuração ─────────────────────────────────────────────────────────────

def load_config(path="config.json") -> dict:
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"❌  Arquivo '{path}' não encontrado. Crie-o com email, password e locations.")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"❌  Erro ao ler '{path}': {e}")
        sys.exit(1)


# ─── Input de datas ───────────────────────────────────────────────────────────

def ask_date(prompt: str, default: str) -> str:
    while True:
        raw = input(f"  {prompt} [{default}]: ").strip()
        val = raw if raw else default
        try:
            datetime.strptime(val, "%Y-%m-%d")
            return val
        except ValueError:
            print("    ⚠  Formato inválido. Use YYYY-MM-DD.")


# ─── Helpers de Excel ─────────────────────────────────────────────────────────

def style_row(ws, row: int, ncols: int, kind: str = "data", alt: bool = False):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = BORDER
        if kind == "header":
            cell.fill = H_FILL
            cell.font = H_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif kind == "total":
            cell.fill = T_FILL
            cell.font = T_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.font = D_FONT
            cell.alignment = Alignment(vertical="center")
            if alt:
                cell.fill = A_FILL


def title_row(ws, row: int, text: str, ncols: int):
    ws.cell(row=row, column=1, value=text).font = TT_FONT
    if ncols > 1:
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=ncols
        )
    ws.row_dimensions[row].height = 22


def write_headers(ws, row: int, headers: list):
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    style_row(ws, row, len(headers), kind="header")
    ws.row_dimensions[row].height = 22


def auto_width(ws, min_w: int = 10, max_w: int = 45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = [len(str(c.value)) for c in col if c.value is not None]
        width = min(max(max(lengths, default=min_w) + 2, min_w), max_w)
        ws.column_dimensions[col_letter].width = width


def safe_val(d, *keys, default="-"):
    """Navega em dicionário aninhado com segurança."""
    for key in keys:
        if not isinstance(d, dict):
            return default
        d = d.get(key, default)
    return d if d != {} else default


def to_list(val) -> list:
    """Garante que o valor retornado seja sempre uma lista."""
    if isinstance(val, list):
        return val
    return []


# ─── Aba: Locais ──────────────────────────────────────────────────────────────

def build_locais(wb: Workbook, locations: list, start_dt: str, end_dt: str):
    ws = wb.create_sheet("Locais")
    headers = ["UUID", "Nome do Local", "Cliente", "Região", "Status"]

    title_row(ws, 1, f"Locais Selecionados  ·  {start_dt}  →  {end_dt}", len(headers))
    write_headers(ws, 2, headers)
    ws.freeze_panes = "A3"

    for i, loc in enumerate(locations, start=3):
        status = "Inativo" if loc.get("deleted_at") else "Ativo"
        ws.cell(row=i, column=1, value=loc.get("uuid", ""))
        ws.cell(row=i, column=2, value=loc.get("name", ""))
        ws.cell(row=i, column=3, value=loc.get("client_name", ""))
        ws.cell(row=i, column=4, value=loc.get("region_name", ""))
        ws.cell(row=i, column=5, value=status)
        style_row(ws, i, len(headers), alt=(i % 2 == 0))

    auto_width(ws)


# ─── Aba: Atividades Resumo ───────────────────────────────────────────────────

def build_atividades_resumo(
    wb: Workbook, counts: dict, eff: dict, start_dt: str, end_dt: str
):
    ws = wb.create_sheet("Atividades Resumo")

    title_row(ws, 1, f"Atividades — Resumo  ·  {start_dt}  →  {end_dt}", 2)
    write_headers(ws, 2, ["Indicador", "Valor"])

    kpis = [
        ("Total de Atividades",      safe_val(counts.get("total"),         "count")),
        ("Completas",                safe_val(counts.get("completed"),      "count")),
        ("Incompletas",              safe_val(counts.get("incomplete"),     "count")),
        ("Perdidas",                 safe_val(counts.get("missed"),         "count")),
        ("Com Eventos",              safe_val(counts.get("events"),         "count")),
        ("Não-Conformidades",        safe_val(counts.get("nonconforming"),  "count")),
        ("", ""),
        ("Check-ins Esperados",      safe_val(eff, "checkinsCount")),
        ("Check-ins Realizados",     safe_val(eff, "checkinsDone")),
        ("Eficiência (%)",           safe_val(eff, "efficiency")),
    ]

    for i, (label, value) in enumerate(kpis, start=3):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        if label:
            style_row(ws, i, 2, alt=(i % 2 == 0))

    auto_width(ws)


# ─── Aba: Atividades por Mês ──────────────────────────────────────────────────

def build_atividades_periodo(
    wb: Workbook, total_month: list, eff_month: list
):
    ws = wb.create_sheet("Atividades por Mês")
    headers = ["Mês", "Total Atividades", "Check-ins Esperados", "Check-ins Feitos", "Eficiência (%)"]

    title_row(ws, 1, "Evolução Mensal de Atividades", len(headers))
    write_headers(ws, 2, headers)
    ws.freeze_panes = "A3"

    eff_idx   = {row.get("date"): row for row in eff_month}
    total_idx = {row.get("date"): row.get("count", 0) for row in total_month}
    all_dates = sorted(set(list(total_idx) + list(eff_idx)))

    for i, d in enumerate(all_dates, start=3):
        e = eff_idx.get(d, {})
        ws.cell(row=i, column=1, value=d)
        ws.cell(row=i, column=2, value=total_idx.get(d, 0))
        ws.cell(row=i, column=3, value=e.get("checkinsCount", ""))
        ws.cell(row=i, column=4, value=e.get("checkinsDone", ""))
        ws.cell(row=i, column=5, value=e.get("efficiency", ""))
        style_row(ws, i, len(headers), alt=(i % 2 == 0))

    # Linha de totais
    n = len(all_dates) + 2
    tr = n + 1
    ws.cell(row=tr, column=1, value="TOTAL / MÉDIA")
    ws.cell(row=tr, column=2, value=f"=SUM(B3:B{n})")
    ws.cell(row=tr, column=3, value=f"=SUM(C3:C{n})")
    ws.cell(row=tr, column=4, value=f"=SUM(D3:D{n})")
    ws.cell(row=tr, column=5, value=f"=IFERROR(AVERAGE(E3:E{n}),\"-\")")
    style_row(ws, tr, len(headers), kind="total")

    auto_width(ws)


# ─── Aba: Checklists ──────────────────────────────────────────────────────────

def build_checklists(
    wb: Workbook,
    count_data: dict,
    eff_data: dict,
    items_gen: list,
    items_nc: list,
):
    ws = wb.create_sheet("Checklists")

    # Bloco 1 — KPIs gerais
    title_row(ws, 1, "Checklists de Atividades — Resumo", 3)
    write_headers(ws, 2, ["Indicador", "Valor"])

    kpis = [
        ("Total de Checklists",         safe_val(count_data, "count")),
        ("Itens Esperados",             safe_val(eff_data,   "checkedsCount")),
        ("Itens Verificados (OK + N/A)",safe_val(eff_data,   "checkedsDone")),
        ("Eficiência (%)",              safe_val(eff_data,   "efficiency")),
    ]
    for i, (label, value) in enumerate(kpis, start=3):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        style_row(ws, i, 2, alt=(i % 2 == 0))

    # Bloco 2 — Itens por nome
    start_bloco2 = len(kpis) + 5
    title_row(ws, start_bloco2, "Itens por Nome (Top ocorrências)", 4)
    headers_items = ["Nome do Item", "Checklist", "Ocorrências", "Tipo"]
    write_headers(ws, start_bloco2 + 1, headers_items)
    ws.freeze_panes = f"A{start_bloco2 + 2}"

    all_items = (
        [(item, "Geral") for item in items_gen]
        + [(item, "Não-Conforme") for item in items_nc]
    )

    for k, (item, tipo) in enumerate(all_items, start=start_bloco2 + 2):
        ws.cell(row=k, column=1, value=item.get("name", ""))
        ws.cell(row=k, column=2, value=item.get("checklist", ""))
        ws.cell(row=k, column=3, value=item.get("count", 0))
        ws.cell(row=k, column=4, value=tipo)
        style_row(ws, k, 4, alt=(k % 2 == 0))

    if not all_items:
        ws.cell(row=start_bloco2 + 2, column=1, value="Nenhum item encontrado no período.")

    auto_width(ws)


# ─── Aba: Justificativas ──────────────────────────────────────────────────────

def build_justificativas(
    wb: Workbook, all_data, inc_data, miss_data
):
    ws = wb.create_sheet("Justificativas")
    headers = ["Justificativa", "Total Geral", "Incompletas", "Perdidas"]

    title_row(ws, 1, "Justificativas de Atividades", len(headers))
    write_headers(ws, 2, headers)
    ws.freeze_panes = "A3"

    def idx(data: list) -> dict:
        return {row["name"]: row["count"] for row in data if "name" in row}

    all_idx  = idx(all_data)
    inc_idx  = idx(inc_data)
    miss_idx = idx(miss_data)
    names    = sorted(set(list(all_idx) + list(inc_idx) + list(miss_idx)))

    for i, name in enumerate(names, start=3):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=all_idx.get(name,  0))
        ws.cell(row=i, column=3, value=inc_idx.get(name,  0))
        ws.cell(row=i, column=4, value=miss_idx.get(name, 0))
        style_row(ws, i, len(headers), alt=(i % 2 == 0))

    if not names:
        ws.cell(row=3, column=1, value="Nenhuma justificativa encontrada no período.")
        return

    n  = len(names) + 2
    tr = n + 1
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=2, value=f"=SUM(B3:B{n})")
    ws.cell(row=tr, column=3, value=f"=SUM(C3:C{n})")
    ws.cell(row=tr, column=4, value=f"=SUM(D3:D{n})")
    style_row(ws, tr, len(headers), kind="total")

    auto_width(ws)


# ─── Aba: Avulsas Perdidas ────────────────────────────────────────────────────

def build_avulsas_perdidas(wb: Workbook, data: list):
    ws = wb.create_sheet("Avulsas Perdidas")
    headers = [
        "ID",
        "Data da Perda",
        "Check-ins Esperados",
        "Check-ins Feitos",
        "Posto",
        "Modelo de Atividade",
        "Local",
        "Região",
        "Cliente",
    ]

    title_row(ws, 1, "Atividades Avulsas Perdidas", len(headers))
    write_headers(ws, 2, headers)
    ws.freeze_panes = "A3"

    for i, row in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=row.get("id", ""))
        ws.cell(row=i, column=2, value=row.get("missed_at", ""))
        ws.cell(row=i, column=3, value=row.get("quantity", 0))
        ws.cell(row=i, column=4, value=row.get("done", 0))
        ws.cell(row=i, column=5, value=(row.get("station") or {}).get("name", ""))
        ws.cell(row=i, column=6, value=(row.get("patrol") or {}).get("name", ""))
        ws.cell(row=i, column=7, value=(row.get("location") or {}).get("name", ""))
        ws.cell(row=i, column=8, value=(row.get("region") or {}).get("name", ""))
        ws.cell(row=i, column=9, value=(row.get("client") or {}).get("name", ""))
        style_row(ws, i, len(headers), alt=(i % 2 == 0))

    if not data:
        ws.cell(row=3, column=1, value="Nenhuma atividade avulsa perdida no período.")

    auto_width(ws)


# ─── Modo Listagem ────────────────────────────────────────────────────────────

def listar_locais(token: str):
    """Lista todos os locais disponíveis na conta e salva CSV auxiliar."""
    print("\n  📍  Buscando todos os locais da conta...\n")
    all_locs = api_get(token, "/filters/locations")

    if not isinstance(all_locs, list) or not all_locs:
        print("  ⚠   Nenhum local retornado pela API.")
        return

    # Separar ativos e inativos
    ativos   = [l for l in all_locs if isinstance(l, dict) and not l.get("deleted_at")]
    inativos = [l for l in all_locs if isinstance(l, dict) and l.get("deleted_at")]

    COL_UUID   = 38
    COL_NOME   = 30
    COL_CLIENT = 25
    COL_REG    = 20

    header = (
        f"  {'UUID':<{COL_UUID}} {'Nome':<{COL_NOME}} "
        f"{'Cliente':<{COL_CLIENT}} {'Região':<{COL_REG}}"
    )
    sep = "  " + "─" * (COL_UUID + COL_NOME + COL_CLIENT + COL_REG + 6)

    print(f"  ✅  LOCAIS ATIVOS ({len(ativos)})")
    print(sep)
    print(header)
    print(sep)
    for loc in sorted(ativos, key=lambda x: x.get("name", "")):
        print(
            f"  {loc.get('uuid',''):<{COL_UUID}} "
            f"{loc.get('name',''):<{COL_NOME}} "
            f"{loc.get('client_name',''):<{COL_CLIENT}} "
            f"{loc.get('region_name',''):<{COL_REG}}"
        )

    if inativos:
        print(f"\n  ⛔  LOCAIS INATIVOS ({len(inativos)})")
        print(sep)
        print(header)
        print(sep)
        for loc in sorted(inativos, key=lambda x: x.get("name", "")):
            print(
                f"  {loc.get('uuid',''):<{COL_UUID}} "
                f"{loc.get('name',''):<{COL_NOME}} "
                f"{loc.get('client_name',''):<{COL_CLIENT}} "
                f"{loc.get('region_name',''):<{COL_REG}}"
            )

    # Salvar CSV auxiliar para facilitar copiar UUIDs
    csv_path = "locais_disponiveis.csv"
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("uuid,nome,cliente,regiao,status\n")
        for loc in sorted(all_locs, key=lambda x: x.get("name", "") if isinstance(x, dict) else ""):
            if not isinstance(loc, dict):
                continue
            status = "inativo" if loc.get("deleted_at") else "ativo"
            nome   = loc.get("name", "").replace(",", ";")
            client = loc.get("client_name", "").replace(",", ";")
            region = loc.get("region_name", "").replace(",", ";")
            f.write(f"{loc.get('uuid','')},{nome},{client},{region},{status}\n")

    print(f"\n  💾  Lista salva em: {csv_path}")
    print("      Copie os UUIDs desejados para o campo 'locations' no config.json.\n")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    SEP = "═" * 56
    print(f"\n{SEP}")
    print("  FindMe Dashboard  ·  Exportador de Indicadores v2.0")
    print(SEP)

    # Config
    cfg       = load_config()
    email     = cfg.get("email", "")
    password  = cfg.get("password", "")
    loc_uuids = cfg.get("locations", [])

    if not email or not password:
        print("❌  Configure 'email' e 'password' no config.json.")
        sys.exit(1)

    # ── Modo listagem: ativado quando locations está vazio no config ──────────
    if not loc_uuids:
        print("\n  ℹ   Nenhum local configurado em config.json.")
        print("      Entrando no modo LISTAGEM para você escolher os locais.\n")
        print("  🔐  Autenticando...")
        token = login(email, password)
        print("  ✔   Token obtido.")
        listar_locais(token)
        sys.exit(0)

    # Período
    print()
    today = date.today().isoformat()
    ago30 = (date.today() - timedelta(days=30)).isoformat()
    start = ask_date("Data início", ago30)
    end   = ask_date("Data fim",    today)
    print()

    # Auth
    print("  🔐  Autenticando...")
    token = login(email, password)
    print("  ✔   Token obtido.\n")

    # Locais
    print("  📍  Buscando locais disponíveis...")
    all_locs = api_get(token, "/filters/locations")
    loc_map  = {
        loc["uuid"]: loc
        for loc in (all_locs if isinstance(all_locs, list) else [])
        if isinstance(loc, dict)
    }
    selected = [loc_map[uid] for uid in loc_uuids if uid in loc_map]
    missing  = [uid for uid in loc_uuids if uid not in loc_map]

    if not selected:
        print("❌  Nenhum local encontrado. Verifique os UUIDs no config.json.")
        sys.exit(1)

    print(f"  ✔   {len(selected)} local(is) confirmado(s):")
    for loc in selected:
        print(f"       • {loc['name']}  ({loc.get('client_name', '')} / {loc.get('region_name', '')})")
    if missing:
        print(f"  ⚠   UUID(s) não encontrado(s): {missing}")

    # Filtro base
    filt = {
        "hiddenInactive": True,
        "locations": loc_uuids,
        "period": [start, end],
    }

    # ── Atividades ────────────────────────────────────────────────────────────
    print("\n  📋  Atividades...")
    counts = {
        "total":         api_post(token, "/activities/count",                filt),
        "completed":     api_post(token, "/activities/count/completed",      filt),
        "incomplete":    api_post(token, "/activities/count/incomplete",     filt),
        "missed":        api_post(token, "/activities/count/missed",         filt),
        "events":        api_post(token, "/activities/count/events",         filt),
        "nonconforming": api_post(token, "/activities/count/non-conforming", filt),
    }
    eff_act     = api_post(token, "/activities/efficiency",              filt)
    total_month = to_list(api_post(token, "/activities/count/period/month",    filt))
    eff_month   = to_list(api_post(token, "/activities/efficiency/period/month", filt))

    # ── Checklists ────────────────────────────────────────────────────────────
    print("  📝  Checklists...")
    count_chk = api_post(token, "/checklists/count",                           filt)
    eff_chk   = api_post(token, "/checklists/efficiency",                      filt)
    items_gen = to_list(api_post(token, "/checklists/count/items/name/general",        filt))
    items_nc  = to_list(api_post(token, "/checklists/count/items/name/non-conforming", filt))

    # ── Justificativas ────────────────────────────────────────────────────────
    print("  📌  Justificativas...")
    just_all  = to_list(api_post(token, "/justifications/count/category",            filt))
    just_inc  = to_list(api_post(token, "/justifications/count/category/incomplete", filt))
    just_miss = to_list(api_post(token, "/justifications/count/category/missed",     filt))

    # ── Avulsas Perdidas ──────────────────────────────────────────────────────
    print("  🔍  Atividades avulsas perdidas...")
    avulsas_raw = api_post(
        token,
        "/reports/missed-single-activities",
        filt,
        params={"page": 1, "limit": 500},
    )
    if isinstance(avulsas_raw, list):
        avulsas = avulsas_raw
    elif isinstance(avulsas_raw, dict):
        avulsas = (
            avulsas_raw.get("data")
            or avulsas_raw.get("rows")
            or avulsas_raw.get("items")
            or []
        )
    else:
        avulsas = []

    # ── Gerar Excel ───────────────────────────────────────────────────────────
    print("\n  📊  Gerando planilha Excel...")
    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão vazia

    build_locais(wb, selected, start, end)
    build_atividades_resumo(wb, counts, eff_act, start, end)
    build_atividades_periodo(wb, total_month, eff_month)
    build_checklists(wb, count_chk, eff_chk, items_gen, items_nc)
    build_justificativas(wb, just_all, just_inc, just_miss)
    build_avulsas_perdidas(wb, avulsas)

    filename = f"findme_{start}_{end}.xlsx"
    wb.save(filename)

    print(f"\n{SEP}")
    print(f"  ✅  Arquivo gerado: {filename}")
    print(f"      Abas criadas:")
    print(f"        1. Locais             — locais selecionados")
    print(f"        2. Atividades Resumo  — KPIs gerais do período")
    print(f"        3. Atividades por Mês — evolução mensal")
    print(f"        4. Checklists         — resumo + itens por nome")
    print(f"        5. Justificativas     — geral / incompletas / perdidas")
    print(f"        6. Avulsas Perdidas   — detalhe de cada atividade")
    print(SEP + "\n")


if __name__ == "__main__":
    main()
