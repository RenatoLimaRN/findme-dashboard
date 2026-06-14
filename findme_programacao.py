#!/usr/bin/env python3
"""
FindMe — Relatório de Atividades por Posto
Gera Excel com Capa Executiva, Atividades detalhadas, Ranking e Grade.

Uso:
    python findme_programacao.py

Lê config.json para credenciais e locais.
Gera: relatorios/YYYY-MM-DD_YYYY-MM-DD/GERAL_YYYY-MM-DD_YYYY-MM-DD.xlsx
         + um arquivo por posto em relatorios/YYYY-MM-DD_YYYY-MM-DD/postos/
"""

import json
import os
import re
import sys
import unicodedata
import requests
from datetime import date, timedelta, datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constantes ───────────────────────────────────────────────────────────────

BASE_DASHBOARD = "https://dashboard-production.findme.id"
BASE_AUTH      = "https://production.api.findme.id"

STATUS = {
    0: "Não iniciada",
    1: "Incompleta",
    2: "Completa",
    4: "Incompleta c/ Justif.",
    5: "Perdida",
}

OP_TYPE = {0: "N/C", 1: "Bombeiro", 2: "Limpeza", 3: "Portaria", 4: "Vigilante"}

DIAS_ORDER = ["Dom", "Seg", "Ter", "Qua", "Qui", "Sex", "Sáb"]

# Paleta de cores
C_DARK   = "1F4E79"
C_BLUE   = "2E75B6"
C_BLUE2  = "4472C4"
C_LIGHT  = "D6E4F0"
C_GREEN  = "1E6B3C"
C_LGREEN = "C6EFCE"
C_RED    = "C0504D"
C_LRED   = "FFCCCC"
C_ORANGE = "C55A11"
C_YELLOW = "FFEB9C"
C_WHITE  = "FFFFFF"
C_GRAY   = "F2F2F2"
C_DARK2  = "2E4057"

THIN   = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NCOLS_CAPA = 8
NCOLS_ATI  = 11
NCOLS_RANK = 8


def mk_fill(hex_color):
    return PatternFill("solid", start_color=hex_color)


def mk_font(bold=False, color=C_WHITE, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)


# ─── Auth & API ───────────────────────────────────────────────────────────────

def login(email: str, password: str) -> str:
    r = requests.post(
        f"{BASE_AUTH}/v3/settings/login",
        json={"email": email, "password": password},
        timeout=30,
    )
    r.raise_for_status()
    data = r.json()
    token = (
        data.get("token") or data.get("access_token")
        or (data.get("data") or {}).get("token")
        or (data.get("data") or {}).get("access_token")
    )
    if not token:
        raise ValueError(f"Token não encontrado: {data}")
    return token


def api_get(token: str, path: str, tentativas: int = 3):
    """GET com retry — a API do dashboard oscila e estoura timeouts curtos."""
    import time
    ultima_exc = None
    for i in range(tentativas):
        try:
            r = requests.get(
                f"{BASE_DASHBOARD}{path}",
                headers={"Authorization": f"Bearer {token}"},
                timeout=90,
            )
            r.raise_for_status()
            return r.json()
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError) as e:
            ultima_exc = e
            if i < tentativas - 1:
                espera = 5 * (i + 1)
                print(f"      ⚠  API lenta ({type(e).__name__}), "
                      f"tentando de novo em {espera}s...")
                time.sleep(espera)
    raise ultima_exc


def fetch_rotinas(token: str, filt: dict, verbose: bool = True) -> tuple:
    """Busca /reports/routines/general com paginação completa e retry.
    Retorna (registros, teve_504): teve_504=True se alguma página falhou por gateway timeout.
    """
    all_rows = []
    had_504  = False
    page = 1
    limit = 50
    MAX_RETRIES = 4
    TIMEOUT = 120

    while True:
        sucesso = False
        for tentativa in range(1, MAX_RETRIES + 1):
            try:
                print(f"    Página {page} (tentativa {tentativa}/{MAX_RETRIES})...")
                r = requests.post(
                    f"{BASE_DASHBOARD}/reports/routines/general",
                    headers={"Authorization": f"Bearer {token}",
                             "Content-Type": "application/json"},
                    json=filt,
                    params={"page": page, "limit": limit},
                    timeout=TIMEOUT,
                    stream=False,
                )
                r.raise_for_status()
                data = r.json()
                sucesso = True
                break
            except requests.exceptions.Timeout:
                print(f"    ⚠  Timeout (tentativa {tentativa}). "
                      f"{'Tentando novamente...' if tentativa < MAX_RETRIES else 'Desistindo.'}")
            except (requests.exceptions.ChunkedEncodingError,
                    requests.exceptions.ConnectionError) as e:
                print(f"    ⚠  Conexão interrompida (tentativa {tentativa}): {type(e).__name__}. "
                      f"{'Tentando novamente...' if tentativa < MAX_RETRIES else 'Desistindo.'}")
            except requests.exceptions.HTTPError as e:
                status_code = e.response.status_code if e.response is not None else 0
                if status_code in (502, 503, 504):
                    had_504 = True
                    print(f"    ⚠  Gateway error {status_code} (tentativa {tentativa}). "
                          f"{'Aguardando 5s e tentando novamente...' if tentativa < MAX_RETRIES else 'Desistindo.'}")
                    if tentativa < MAX_RETRIES:
                        import time; time.sleep(5)
                else:
                    print(f"    ⚠  HTTP {status_code} (tentativa {tentativa}): {e}")
                    break
            except Exception as e:
                print(f"    ⚠  Erro inesperado (tentativa {tentativa}): {e}")
                break

        if not sucesso:
            print(f"    ❌  Página {page} falhou. Parando.")
            break

        if isinstance(data, list):
            rows = data
        elif isinstance(data, dict):
            rows = data.get("data") or data.get("rows") or data.get("items") or []
        else:
            rows = []

        if not rows:
            print(f"    ✔  Página {page} vazia — fim da paginação.")
            break

        all_rows.extend(rows)
        print(f"    ✔  Página {page}: {len(rows)} registros (total: {len(all_rows)})")

        if len(rows) < limit:
            break
        page += 1

    return all_rows, had_504


def fetch_rotinas_por_dia(token: str, filt_base: dict,
                          start: str, end: str, nome: str = "") -> list:
    """Fallback: busca 1 dia por vez para locais que dão 504 no período completo.
    Divide o período em dias e combina os resultados — nunca sobrecarrega a API.
    """
    import time
    d_ini = datetime.strptime(start, "%Y-%m-%d")
    d_fim = datetime.strptime(end,   "%Y-%m-%d")
    total_dias = (d_fim - d_ini).days + 1
    all_rows = []
    d = d_ini
    dia_n = 0

    while d <= d_fim:
        dia_n += 1
        day = d.strftime("%Y-%m-%d")
        filt_day = dict(filt_base)
        filt_day["period"] = [day, day]
        rows, _ = fetch_rotinas(token, filt_day)
        if rows:
            all_rows.extend(rows)
            print(f"      [{dia_n}/{total_dias}] {day}: {len(rows)} atividades")
        else:
            print(f"      [{dia_n}/{total_dias}] {day}: vazio")
        time.sleep(0.3)   # respeita a API entre dias
        d += timedelta(days=1)

    return all_rows


def buscar_modelos_historico(token: str, loc: dict, start: str) -> list:
    """Para locais sem atividades no período, busca os 60 dias anteriores ao início.
    Extrai combinações únicas de posto + modelo para mostrar o que está configurado.
    Retorna lista de dicts: {posto, op_tipo, modelo}
    """
    d_fim = datetime.strptime(start, "%Y-%m-%d") - timedelta(days=1)
    d_ini = d_fim - timedelta(days=60)
    filt = {
        "hiddenInactive": True,
        "locations": [loc["uuid"]],
        "period": [d_ini.strftime("%Y-%m-%d"), d_fim.strftime("%Y-%m-%d")],
    }
    rows, _ = fetch_rotinas(token, filt)
    if not rows:
        return []

    vistos  = set()
    modelos = []
    for row in rows:
        station = row.get("station") or {}
        patrol  = row.get("patrol")  or {}
        posto   = station.get("name", "-")
        op_tipo = OP_TYPE.get(station.get("operation_type", 0), "N/C")
        modelo  = patrol.get("name", "-")
        chave   = (posto, modelo)
        if chave not in vistos:
            vistos.add(chave)
            modelos.append({"posto": posto, "op_tipo": op_tipo, "modelo": modelo})

    return sorted(modelos, key=lambda x: (x["op_tipo"], x["posto"], x["modelo"]))


# ─── Helpers ──────────────────────────────────────────────────────────────────

def load_config(path="config.json") -> dict:
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"❌  '{path}' não encontrado.")
        sys.exit(1)


def ask_date(prompt: str, default: str) -> str:
    while True:
        raw = input(f"  {prompt} [{default}]: ").strip()
        val = raw if raw else default
        try:
            datetime.strptime(val, "%Y-%m-%d")
            return val
        except ValueError:
            print("    ⚠  Use YYYY-MM-DD.")


def parse_dt(s: str):
    if not s:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S.%f"):
        try:
            return datetime.strptime(s[:19], fmt[:len(fmt)])
        except Exception:
            pass
    return None


def hora_fmt(dt) -> str:
    return dt.strftime("%H:%M") if dt else "-"


def data_fmt(dt) -> str:
    return dt.strftime("%d/%m/%Y") if dt else "-"


def dia_semana(dt) -> str:
    if not dt:
        return "-"
    dias = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    return dias[dt.weekday()]


def turno(dt) -> str:
    if not dt:
        return "-"
    h = dt.hour
    if 0 <= h < 6:   return "Madrugada (00h-06h)"
    if 6 <= h < 12:  return "Manhã (06h-12h)"
    if 12 <= h < 18: return "Tarde (12h-18h)"
    return "Noite (18h-00h)"


def duracao(dt_ini, dt_fim) -> str:
    if not dt_ini or not dt_fim:
        return "-"
    delta = dt_fim - dt_ini
    if delta.total_seconds() <= 0:
        return "-"
    total_seg = int(delta.total_seconds())
    horas   = total_seg // 3600
    minutos = (total_seg % 3600) // 60
    segundos = total_seg % 60
    if horas > 0:
        return f"{horas}h {minutos}min {segundos}s"
    if minutos > 0:
        return f"{minutos} minutos e {segundos} segundos"
    return f"{segundos} segundos"


def barra_progresso(pct: float, total: int = 10) -> str:
    filled = round(pct * total / 100)
    filled = max(0, min(total, filled))
    return "█" * filled + "░" * (total - filled)


def classificar_status(status_int: int) -> str:
    """Retorna: OK / Parcial / Não Feita"""
    if status_int == 2:
        return "OK"
    if status_int in (1, 4):
        return "Parcial"
    return "Não Feita"  # 0 e 5


def pct_eficiencia(ok: int, parcial: int, total: int) -> float:
    """% = (OK + Parcial * 0.5) / Total * 100"""
    if total == 0:
        return 0.0
    return round((ok + parcial * 0.5) / total * 100, 1)


def extrair_justificativa(row: dict) -> str:
    """Extrai texto da justificativa da atividade.

    Estrutura real (API Dashboard v2.0, doc):
      justifications: [
        {id, justification_category_id, justification_category_name,
         justification_description, created_at, user}
      ]

    Retorna o texto mais informativo. Se houver várias justificativas,
    junta com " | ".
    """
    just = row.get("justifications") or []
    if isinstance(just, dict):
        just = [just]
    if not isinstance(just, list):
        return "-"
    textos = []
    for j in just:
        if not isinstance(j, dict):
            continue
        # campos corretos da API + fallbacks pra resiliência
        cat = (j.get("justification_category_name")
               or j.get("category_name") or j.get("name") or "").strip()
        desc = (j.get("justification_description")
                or j.get("description") or j.get("reason")
                or j.get("text") or "").strip()
        if cat and desc:
            textos.append(f"{cat} — {desc}")
        elif cat:
            textos.append(cat)
        elif desc:
            textos.append(desc)
    return " | ".join(textos) if textos else "-"


# ─── Injeção de avulsas configuradas ─────────────────────────────────────────

DIA_MAP = {"Dom": 6, "Seg": 0, "Ter": 1, "Qua": 2, "Qui": 3, "Sex": 4, "Sab": 5}


def injetar_avulsas_config(dados: dict, avulsas_cfg: list,
                           start: str, end: str) -> int:
    """
    Para cada local/posto/atividade definido em avulsas_cfg (config.json),
    verifica se apareceu nos dados da API no período. Se não apareceu,
    injeta um registro 'Não iniciada'.
    Retorna o número de registros injetados.
    """
    if not avulsas_cfg:
        return 0

    d_ini = datetime.strptime(start, "%Y-%m-%d")
    d_fim = datetime.strptime(end,   "%Y-%m-%d")

    def _norm_txt(s: str) -> str:
        """lower + sem acentos + espaços colapsados — o postos/*.json costuma
        vir sem acento ('CONDOMINIO') e a API devolve com ('CONDOMÍNIO');
        sem normalizar, a avulsa feita nunca casa e vira falso 'não feita'
        num local fantasma duplicado."""
        s = unicodedata.normalize("NFD", s or "")
        s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
        return re.sub(r"\s+", " ", s).strip().lower()

    # Nome real (da API) por nome normalizado — pra injetar no grupo certo
    locais_reais = {_norm_txt(nome): nome for nome in dados["por_local"]}

    # Índice rápido: (local_norm, modelo_norm, data_dd/mm/yyyy) → count.
    # Posto fica de fora de propósito: o nome do posto no postos/*.json
    # raramente bate letra a letra com o station da API.
    from collections import Counter
    existentes = Counter()
    for rec in dados["records"]:
        chave = (
            _norm_txt(rec["local"]),
            _norm_txt(rec["modelo"]),
            rec["data"],
        )
        existentes[chave] += 1

    n_inj = 0
    d = d_ini
    while d <= d_fim:
        dia_python = d.weekday()   # 0=Seg … 6=Dom
        dia_abbr   = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"][dia_python]
        date_str   = d.strftime("%d/%m/%Y")

        for cfg_local in avulsas_cfg:
            # injeta no grupo do local REAL da API (com acento), se existir
            local_nome = locais_reais.get(
                _norm_txt(cfg_local["local"]), cfg_local["local"])

            for cfg_posto in cfg_local.get("postos", []):
                posto_nome = cfg_posto["posto"]
                op_tipo    = cfg_posto.get("op_tipo", "Limpeza")

                for ativ in cfg_posto.get("atividades", []):
                    if dia_abbr not in ativ.get("dias", []):
                        continue
                    modelo = ativ["modelo"]
                    vezes  = ativ.get("vezes", 1)
                    if not isinstance(vezes, int) or vezes <= 0:
                        continue  # vezes=0 → sob demanda, não cobra

                    chave = (_norm_txt(local_nome),
                             _norm_txt(modelo),
                             date_str)
                    # consome o crédito: a mesma execução não pode satisfazer
                    # duas expectativas do mesmo modelo no mesmo dia
                    ja_feitas = min(vezes, existentes.get(chave, 0))
                    if ja_feitas:
                        existentes[chave] -= ja_feitas
                    faltam = vezes - ja_feitas

                    for _ in range(faltam):
                        rec = {
                            "local":         local_nome,
                            "regiao":        "-",
                            "cliente":       "-",
                            "posto":         posto_nome,
                            "op_tipo":       op_tipo,
                            "modelo":        modelo,
                            "avulsa":        True,
                            # esperada no postos/*.json mas o FindMe nem criou —
                            # diferente de "Não iniciada" (criada e não executada)
                            "status_int":    -1,
                            "status_label":  "Esperada — Não Registrada",
                            "status_class":  "Não Feita",
                            "data":          date_str,
                            "hora":          "-",
                            "turno":         "-",
                            "hora_fim_prog": "-",
                            "iniciada":      "-",
                            "finalizada":    "-",
                            "duracao":       "-",
                            "dia_semana":    dia_abbr,
                            "pontos_ok":     0,
                            "pontos_total":  0,
                            "justificativa": "-",
                            "dt_sort":       d,
                        }
                        dados["records"].append(rec)
                        if local_nome not in dados["por_local"]:
                            dados["por_local"][local_nome] = {
                                "ok": 0, "parcial": 0, "nao_feita": 0,
                                "total": 0, "pct": 0.0, "records": []
                            }
                        dl = dados["por_local"][local_nome]
                        dl["nao_feita"] += 1
                        dl["total"]     += 1
                        dl["pct"] = pct_eficiencia(dl["ok"], dl["parcial"], dl["total"])
                        dl["records"].append(rec)
                        n_inj += 1

        d += timedelta(days=1)

    return n_inj


# ─── Processamento ────────────────────────────────────────────────────────────

def processar(rows: list) -> dict:
    """
    Retorna dict com:
      - records:   lista completa de atividades com campos calculados
      - por_local: {nome_local → {ok, parcial, nao_feita, total, pct, records}}
    """
    records = []

    for row in rows:
        if not isinstance(row, dict):
            continue

        station    = row.get("station") or {}
        op_type    = station.get("operation_type", 0)
        is_single  = row.get("single", False)
        status_int = row.get("status", 0)

        dt_prog_ini = parse_dt(row.get("to_be_started_at"))
        dt_prog_fim = parse_dt(row.get("to_be_finished_until"))
        dt_real_ini = parse_dt(row.get("started_at"))
        dt_real_fim = parse_dt(row.get("finished_at"))

        rec = {
            "local":         (row.get("location") or {}).get("name", "-"),
            "regiao":        (row.get("region") or {}).get("name", "-"),
            "cliente":       (row.get("client") or {}).get("name", "-"),
            "posto":         station.get("name", "-"),
            "op_tipo":       OP_TYPE.get(op_type, str(op_type)),
            "modelo":        (row.get("patrol") or {}).get("name", "-"),
            "avulsa":        is_single,
            "status_int":    status_int,
            "status_label":  STATUS.get(status_int, "-"),
            "status_class":  classificar_status(status_int),
            "data":          data_fmt(dt_prog_ini),
            "hora":          hora_fmt(dt_prog_ini),
            "turno":         turno(dt_prog_ini),
            "hora_fim_prog": hora_fmt(dt_prog_fim),
            "iniciada":      hora_fmt(dt_real_ini),
            "finalizada":    hora_fmt(dt_real_fim),
            "duracao":       duracao(dt_real_ini, dt_real_fim),
            "dia_semana":    dia_semana(dt_prog_ini),
            "pontos_ok":     row.get("checkins_done", 0),
            "pontos_total":  row.get("checkins_count", 0),
            "justificativa": extrair_justificativa(row),
            "dt_sort":       dt_prog_ini or datetime.min,
        }
        records.append(rec)

    # Agrupa por local
    por_local = defaultdict(lambda: {
        "ok": 0, "parcial": 0, "nao_feita": 0,
        "total": 0, "pct": 0.0, "records": []
    })

    for rec in records:
        d = por_local[rec["local"]]
        d["total"] += 1
        d["records"].append(rec)
        sc = rec["status_class"]
        if sc == "OK":
            d["ok"] += 1
        elif sc == "Parcial":
            d["parcial"] += 1
        else:
            d["nao_feita"] += 1

    for d in por_local.values():
        d["pct"] = pct_eficiencia(d["ok"], d["parcial"], d["total"])

    return {"records": records, "por_local": dict(por_local)}


# ─── Excel helpers ────────────────────────────────────────────────────────────

def cel(ws, row, col, value=None, fill=None, font=None,
        align=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = mk_fill(fill)
    if font:
        c.font = font
    if align:
        c.alignment = align
    if border:
        c.border = BORDER
    return c


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)


def cor_eficiencia(pct: float):
    """Retorna (fill_hex, font_hex) baseado na eficiência."""
    if pct >= 90:
        return C_LGREEN, C_GREEN
    if pct >= 70:
        return C_YELLOW, "7F6000"
    return C_LRED, C_RED


def status_fill(status_label: str, status_class: str = ""):
    """Retorna (fill_hex, font_hex) pelo rótulo real do status.

    Distingue "Perdida" de "Não iniciada" e destaca
    "Esperada — Não Registrada" (avulsa do postos/ que o sistema nem criou),
    em vez de pintar tudo como um "Não Feita" genérico.
    """
    por_label = {
        "Completa":                  (C_LGREEN, C_GREEN),
        "Incompleta":                (C_YELLOW, "7F6000"),
        "Incompleta c/ Justif.":     (C_YELLOW, "7F6000"),
        "Não iniciada":              (C_LRED,   C_RED),
        "Perdida":                   ("F4A7A3", "7B1F1C"),
        "Esperada — Não Registrada": ("E57373", "FFFFFF"),
    }
    if status_label in por_label:
        return por_label[status_label]
    return {
        "OK":        (C_LGREEN, C_GREEN),
        "Parcial":   (C_YELLOW, "7F6000"),
        "Não Feita": (C_LRED,   C_RED),
    }.get(status_class, (C_WHITE, "000000"))


# ─── ABA: Capa Executiva ──────────────────────────────────────────────────────

def build_capa(wb, dados: dict, start: str, end: str):
    ws = wb.create_sheet("Capa Executiva")
    ws.sheet_view.showGridLines = False

    por_local = dados["por_local"]
    records   = dados["records"]
    total_ati = len(records)

    ok_total      = sum(d["ok"]       for d in por_local.values())
    parcial_total = sum(d["parcial"]  for d in por_local.values())
    nao_total     = sum(d["nao_feita"] for d in por_local.values())
    pct_geral     = pct_eficiencia(ok_total, parcial_total, total_ati)
    n_locais      = len(por_local)
    n_criticos    = sum(1 for d in por_local.values() if d["pct"] < 70)

    gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
    dt_start  = datetime.strptime(start, "%Y-%m-%d").strftime("%d/%m/%Y")
    dt_end    = datetime.strptime(end,   "%Y-%m-%d").strftime("%d/%m/%Y")
    periodo   = f"{dt_start}  →  {dt_end}"

    # ── Título ──────────────────────────────────────────────────────────────
    merge(ws, 1, 1, 1, NCOLS_CAPA)
    c = cel(ws, 1, 1, "📊 PAINEL DE ATIVIDADES — FindMe", fill=C_DARK)
    c.font = Font(name="Arial", bold=True, size=18, color=C_WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    merge(ws, 2, 1, 2, NCOLS_CAPA)
    c = cel(ws, 2, 1, f"Período: {periodo}   •   Gerado em {gerado_em}",
            fill=C_DARK2)
    c.font = Font(name="Arial", italic=True, size=10, color="D0D0D0")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    # ── KPI Cards ────────────────────────────────────────────────────────────
    # linha 4 = valor grande, linha 5 = label
    pct_cor = C_RED if pct_geral < 70 else (C_BLUE2 if pct_geral < 90 else C_GREEN)
    kpis = [
        (f"{pct_geral}%",  "% Eficiência Geral",       pct_cor,                                          (1, 2)),
        (n_locais,          "Locais Monitorados",        C_BLUE2,                                          (3, 4)),
        (n_criticos,        "Locais Críticos (<70%)",    C_RED if n_criticos > 0 else C_BLUE2,             (5, 6)),
        (nao_total,         "Atividades Não Feitas",     C_RED if nao_total > 0 else C_GREEN,              (7, 8)),
    ]

    for val, label, cor, (c1, c2) in kpis:
        merge(ws, 4, c1, 4, c2)
        merge(ws, 5, c1, 5, c2)
        v = ws.cell(row=4, column=c1, value=val)
        v.fill = mk_fill(cor)
        v.font = Font(name="Arial", bold=True, size=28, color=C_WHITE)
        v.alignment = Alignment(horizontal="center", vertical="center")
        v.border = BORDER
        l = ws.cell(row=5, column=c1, value=label)
        l.fill = mk_fill(cor)
        l.font = Font(name="Arial", size=9, color="E8E8E8")
        l.alignment = Alignment(horizontal="center", vertical="center")
        l.border = BORDER

    ws.row_dimensions[4].height = 48
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 12

    # ── Programadas × Avulsas, status a status ───────────────────────────────
    def contar(recs: list) -> dict:
        n = {"feitas": 0, "parciais": 0, "nao_iniciadas": 0,
             "perdidas": 0, "esperadas_nr": 0}
        for r in recs:
            lbl = r["status_label"]
            if lbl == "Completa":
                n["feitas"] += 1
            elif lbl.startswith("Incompleta"):
                n["parciais"] += 1
            elif lbl == "Perdida":
                n["perdidas"] += 1
            elif lbl == "Esperada — Não Registrada":
                n["esperadas_nr"] += 1
            else:
                n["nao_iniciadas"] += 1
        return n

    prog_recs = [r for r in records if not r["avulsa"]]
    avu_recs  = [r for r in records if r["avulsa"]]

    merge(ws, 7, 1, 7, NCOLS_CAPA)
    c = cel(ws, 7, 1, "📋  PROGRAMADAS × AVULSAS — visão por status", fill=C_DARK2)
    c.font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[7].height = 22

    hdrs = ["Origem", "✓ Feitas", "⚠ Parciais", "✗ Não iniciadas",
            "⊘ Perdidas", "❌ Esp. não registr.", "Total", "% Efic."]
    hfills = [C_DARK, C_GREEN, C_ORANGE, C_RED, "7B1F1C", "B71C1C", C_BLUE2, C_DARK]
    for col, (h, hf) in enumerate(zip(hdrs, hfills), 1):
        c = cel(ws, 8, col, h, fill=hf)
        c.font = mk_font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 18

    for i, (nome_origem, recs) in enumerate(
            [("Programadas", prog_recs), ("Avulsas", avu_recs)]):
        n = contar(recs)
        total = len(recs)
        pct = pct_eficiencia(n["feitas"], n["parciais"], total)
        row = 9 + i
        esp_nr = n["esperadas_nr"] if nome_origem == "Avulsas" else "—"
        vals = [nome_origem, n["feitas"], n["parciais"], n["nao_iniciadas"],
                n["perdidas"], esp_nr, total, f"{pct}%"]
        bg = C_GRAY if i % 2 == 0 else C_WHITE
        for col, val in enumerate(vals, 1):
            c = cel(ws, row, col, val, fill=bg)
            c.font = Font(name="Arial", size=10, color="000000",
                          bold=(col in (1, 8)))
            c.alignment = Alignment(
                horizontal="left" if col == 1 else "center",
                vertical="center", indent=1 if col == 1 else 0)
        fc, ft = cor_eficiencia(pct)
        c8 = ws.cell(row=row, column=8)
        c8.fill = mk_fill(fc)
        c8.font = Font(name="Arial", bold=True, size=10, color=ft)
        ws.row_dimensions[row].height = 18

    ws.row_dimensions[11].height = 12

    # ── Tabela de Postos ─────────────────────────────────────────────────────
    def tabela_postos(row_start: int, titulo: str, locais: list,
                      cor_titulo: str) -> int:
        merge(ws, row_start, 1, row_start, NCOLS_CAPA)
        c = cel(ws, row_start, 1, titulo, fill=cor_titulo)
        c.font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row_start].height = 24

        hdrs = ["#", "Local / Posto", "✓ OK",
                "⚠ Parcial", "✗ Não Feita", "Total", "% Efic.", "Progresso"]
        hfills = [C_DARK, C_DARK, C_GREEN, C_ORANGE, C_RED, C_BLUE2, C_DARK, C_DARK]

        for col, (h, hf) in enumerate(zip(hdrs, hfills), 1):
            c = cel(ws, row_start + 1, col, h, fill=hf)
            c.font = mk_font(bold=True, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_start + 1].height = 18

        for i, (nome, d) in enumerate(locais):
            row = row_start + 2 + i
            bg  = C_GRAY if i % 2 == 0 else C_WHITE
            pct = d["pct"]
            fc, ft = cor_eficiencia(pct)

            vals = [i + 1, nome, d["ok"], d["parcial"],
                    d["nao_feita"], d["total"], f"{pct}%",
                    barra_progresso(pct)]

            for col, val in enumerate(vals, 1):
                c = cel(ws, row, col, val, fill=bg)
                c.font = Font(name="Arial", size=9, color="000000")
                c.alignment = Alignment(
                    horizontal="left" if col == 2 else "center",
                    vertical="center",
                    indent=1 if col == 2 else 0,
                )
                ws.row_dimensions[row].height = 16

            # % eficiência colorida
            c7 = ws.cell(row=row, column=7)
            c7.fill = mk_fill(fc)
            c7.font = Font(name="Arial", bold=True, size=9, color=ft)

        return row_start + 2 + len(locais)

    locais_sorted = sorted(por_local.items(), key=lambda x: x[1]["pct"])
    criticos = [(n, d) for n, d in locais_sorted if d["pct"] < 70]
    otimos   = [(n, d) for n, d in reversed(locais_sorted) if d["pct"] >= 90]
    todos    = list(reversed(locais_sorted))

    row_cur = 12  # depois da tabela Programadas × Avulsas

    if criticos:
        row_cur = tabela_postos(
            row_cur,
            f"🚨  LOCAIS CRÍTICOS — AÇÃO URGENTE  ({len(criticos)} locais abaixo de 70%)",
            criticos, C_RED,
        )
        row_cur += 1

    if otimos:
        row_cur = tabela_postos(
            row_cur,
            f"✓  LOCAIS COM ÓTIMO DESEMPENHO  ({len(otimos)} locais ≥ 90%)",
            otimos, C_GREEN,
        )
        row_cur += 1

    tabela_postos(
        row_cur,
        f"📋  TODOS OS LOCAIS  ({len(todos)})",
        todos, C_DARK,
    )

    # Larguras
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14


# ─── ABA: Atividades ──────────────────────────────────────────────────────────

def build_atividades(wb, dados: dict, start: str, end: str,
                     todos_locais: list = None,
                     modelos_historicos: dict = None):
    """
    todos_locais:       lista de nomes de todos os locais selecionados,
                        inclusive os que retornaram 0 atividades da API.
    modelos_historicos: dict nome_local → lista de {posto, op_tipo, modelo}
                        para locais sem dados no período (buscados nos 60 dias anteriores).
    """
    ws = wb.create_sheet("Atividades")
    ws.sheet_view.showGridLines = False

    headers = [
        "Data", "Hora", "Turno", "Tipo / Modelo",
        "Status", "Iniciada", "Finalizada", "Duração",
        "Pts OK", "Pts Total", "Justificativa",
    ]

    # Linha de cabeçalho fixo
    for col, h in enumerate(headers, 1):
        c = cel(ws, 1, col, h, fill=C_BLUE2)
        c.font = mk_font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    row_cur = 2
    por_local = dados["por_local"]

    # Garante que todos os locais selecionados apareçam, mesmo sem atividades
    nomes_ordenados = sorted(todos_locais) if todos_locais else sorted(por_local.keys())

    for local_nome in nomes_ordenados:
        d = por_local.get(local_nome)

        if not d or not d["records"]:
            # Cabeçalho do local — vermelho/cinza para indicar ausência
            merge(ws, row_cur, 1, row_cur, NCOLS_ATI)
            c = cel(ws, row_cur, 1,
                    f"  📍  {local_nome}   —   Sem atividades registradas no período",
                    fill="6D6D6D")
            c.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_cur].height = 20
            row_cur += 1

            hist = (modelos_historicos or {}).get(local_nome, [])
            if hist:
                # Sub-cabeçalho informando origem dos dados
                merge(ws, row_cur, 1, row_cur, NCOLS_ATI)
                c = cel(ws, row_cur, 1,
                        "     📋  Atividades configuradas (histórico 60 dias anteriores ao período) "
                        "— 0 execuções registradas neste período",
                        fill="EBF3FB")
                c.font = Font(name="Arial", italic=True, size=9, color=C_BLUE)
                c.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[row_cur].height = 16
                row_cur += 1

                # Mini-cabeçalho da tabela de modelos
                hdrs = ["Tipo", "Posto", "Modelo de Atividade",
                        "", "", "", "", "", "", "", ""]
                for col, h in enumerate(hdrs, 1):
                    c = cel(ws, row_cur, col, h,
                            fill="D6E4F0" if h else C_GRAY)
                    c.font = Font(name="Arial", bold=True, size=8, color="1F4E79")
                    c.alignment = Alignment(horizontal="center" if col == 1
                                            else "left", vertical="center")
                ws.row_dimensions[row_cur].height = 14
                row_cur += 1

                # Linhas de cada modelo
                for j, m in enumerate(hist):
                    bg = "F2F8FF" if j % 2 == 0 else C_WHITE
                    vals = [m["op_tipo"], m["posto"], m["modelo"],
                            "", "", "", "", "", "", "", ""]
                    for col, val in enumerate(vals, 1):
                        c = cel(ws, row_cur, col, val, fill=bg)
                        c.font = Font(name="Arial", size=8,
                                      color="888888" if col > 3 else "333333",
                                      italic=(col > 3))
                        c.alignment = Alignment(
                            horizontal="center" if col == 1 else "left",
                            vertical="center", indent=1 if col in (2, 3) else 0)
                    ws.row_dimensions[row_cur].height = 14
                    row_cur += 1
            else:
                # Sem histórico — aviso simples
                merge(ws, row_cur, 1, row_cur, NCOLS_ATI)
                c = cel(ws, row_cur, 1,
                        "     ⚠  Sem atividades no período e sem histórico anterior "
                        "(local possivelmente novo ou inativo)",
                        fill="F2F2F2")
                c.font = Font(name="Arial", italic=True, size=9, color="888888")
                c.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[row_cur].height = 16
                row_cur += 1
            continue

        recs = sorted(d["records"], key=lambda r: r["dt_sort"])

        # Conta avulsas separado
        n_avulsas = sum(1 for r in recs if r["avulsa"])
        n_prog    = len(recs) - n_avulsas

        # Quebra real por status (não amassa Perdida com Não iniciada)
        n_ni   = sum(1 for r in recs if r["status_label"] == "Não iniciada")
        n_perd = sum(1 for r in recs if r["status_label"] == "Perdida")
        n_enr  = sum(1 for r in recs
                     if r["status_label"] == "Esperada — Não Registrada")

        # Separador de local
        merge(ws, row_cur, 1, row_cur, NCOLS_ATI)
        partes = [f"✓ {d['ok']} Feitas", f"⚠ {d['parcial']} Parciais"]
        if n_ni:
            partes.append(f"✗ {n_ni} Não iniciadas")
        if n_perd:
            partes.append(f"⊘ {n_perd} Perdidas")
        if n_enr:
            partes.append(f"❌ {n_enr} Esperadas não registradas")
        resumo = (f"  📍  {local_nome}   —   " + "   ".join(partes)
                  + f"   |   {d['pct']}% eficiência"
                  + f"   |   {n_prog} programada(s)"
                  + (f" + 🔔 {n_avulsas} avulsa(s)" if n_avulsas else ""))
        c = cel(ws, row_cur, 1, resumo, fill=C_BLUE2)
        c.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_cur].height = 20
        row_cur += 1

        for i, rec in enumerate(recs):
            sc     = rec["status_class"]
            sf, st = status_fill(rec["status_label"], sc)
            is_avulsa = rec["avulsa"]

            # Avulsas têm fundo levemente diferente para destacar
            if is_avulsa:
                row_bg = "FFF3E0" if i % 2 == 0 else "FFF8F0"
            else:
                row_bg = C_GRAY if i % 2 == 0 else C_WHITE

            modelo_label = (f"[AVULSA] {rec['modelo']}"
                            if is_avulsa else rec["modelo"])

            vals = [
                rec["data"], rec["hora"], rec["turno"], modelo_label,
                rec["status_label"],
                rec["iniciada"], rec["finalizada"], rec["duracao"],
                rec["pontos_ok"], rec["pontos_total"],
                rec["justificativa"],
            ]

            for col, val in enumerate(vals, 1):
                c = cel(ws, row_cur, col,
                        val, fill=row_bg if col != 5 else sf)
                if col == 5:
                    c.font = Font(name="Arial", bold=True, size=9, color=st)
                elif col == 4 and is_avulsa:
                    # Modelo com [AVULSA] em laranja
                    c.font = Font(name="Arial", size=9, color=C_ORANGE, bold=True)
                else:
                    c.font = Font(name="Arial", size=9, color="000000")
                c.alignment = Alignment(
                    horizontal="center" if col in (1, 2, 6, 7, 8, 9, 10) else "left",
                    vertical="center",
                    indent=1 if col in (3, 4, 11) else 0,
                )
            ws.row_dimensions[row_cur].height = 15
            row_cur += 1

    # Larguras
    widths = [12, 8, 22, 34, 22, 10, 10, 20, 8, 8, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─── ABA: Ranking ─────────────────────────────────────────────────────────────

def build_ranking(wb, dados: dict, start: str, end: str):
    ws = wb.create_sheet("🏆 Ranking")
    ws.sheet_view.showGridLines = False

    por_local = dados["por_local"]
    records   = dados["records"]
    total_ati = len(records)

    ok_total  = sum(d["ok"]       for d in por_local.values())
    par_total = sum(d["parcial"]  for d in por_local.values())
    nao_total = sum(d["nao_feita"] for d in por_local.values())
    pct_geral = pct_eficiencia(ok_total, par_total, total_ati)
    n_otimos  = sum(1 for d in por_local.values() if d["pct"] >= 90)
    n_crit    = sum(1 for d in por_local.values() if d["pct"] < 70)

    gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
    dt_start  = datetime.strptime(start, "%Y-%m-%d").strftime("%d/%m/%Y")
    dt_end    = datetime.strptime(end,   "%Y-%m-%d").strftime("%d/%m/%Y")
    periodo   = f"{dt_start}  →  {dt_end}"

    # Título
    merge(ws, 1, 1, 1, NCOLS_RANK)
    c = cel(ws, 1, 1, "🏆 RANKING GERAL DA OPERAÇÃO", fill=C_DARK)
    c.font = Font(name="Arial", bold=True, size=14, color=C_WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    merge(ws, 2, 1, 2, NCOLS_RANK)
    c = cel(ws, 2, 1,
            f"Período: {periodo}  |  Gerado em {gerado_em}",
            fill=C_DARK2)
    c.font = Font(name="Arial", italic=True, size=9, color="D0D0D0")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 10

    # Indicadores gerais
    merge(ws, 4, 1, 4, NCOLS_RANK)
    c = cel(ws, 4, 1, "📊 INDICADORES GERAIS DA OPERAÇÃO", fill=C_BLUE2)
    c.font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 20

    ind_hdrs = ["Locais", "Atividades", "% Eficiência",
                "✓ Ótimos (≥90%)", "⚠ Críticos (<70%)",
                "✗ Não Feitas", "✓ OK", "⚠ Parciais"]
    ind_vals = [len(por_local), total_ati, f"{pct_geral}%",
                n_otimos, n_crit, nao_total, ok_total, par_total]
    ind_hfills = [C_BLUE2, C_BLUE2, C_BLUE2,
                  C_GREEN, C_RED, C_RED, C_GREEN, C_ORANGE]
    ind_vfills = [C_LIGHT, C_LIGHT,
                  C_LGREEN if pct_geral >= 70 else C_LRED,
                  C_LGREEN, C_LRED, C_LRED, C_LGREEN, C_YELLOW]

    for col, (h, v, fh, fv) in enumerate(
            zip(ind_hdrs, ind_vals, ind_hfills, ind_vfills), 1):
        c = cel(ws, 5, col, h, fill=fh)
        c.font = mk_font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c2 = cel(ws, 6, col, v, fill=fv)
        c2.font = Font(name="Arial", bold=True, size=14, color="000000")
        c2.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 32
    ws.row_dimensions[7].height = 10

    locais_sorted = sorted(por_local.items(), key=lambda x: -x[1]["pct"])
    MEDALHAS = ["🥇", "🥈", "🥉", "4º", "5º", "6º", "7º", "8º", "9º", "10º"]

    def secao_ranking(row_start: int, titulo: str, cor: str,
                      locais: list, medalhas: bool = False) -> int:
        merge(ws, row_start, 1, row_start, NCOLS_RANK)
        c = cel(ws, row_start, 1, titulo, fill=cor)
        c.font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row_start].height = 22

        hdrs = ["#", "Local", "Total",
                "✓ OK", "⚠ Parcial", "✗ Não Feitas",
                "% Efic.", "Progresso"]
        for col, h in enumerate(hdrs, 1):
            c = cel(ws, row_start + 1, col, h, fill=C_DARK)
            c.font = mk_font(bold=True, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_start + 1].height = 16

        for i, (nome, d) in enumerate(locais):
            row = row_start + 2 + i
            bg  = C_LIGHT if i % 2 == 0 else C_WHITE
            pct = d["pct"]
            fc, ft = cor_eficiencia(pct)
            rank_label = MEDALHAS[i] if medalhas and i < len(MEDALHAS) else str(i + 1)

            vals = [rank_label, nome, d["total"],
                    d["ok"], d["parcial"], d["nao_feita"],
                    f"{pct}%", barra_progresso(pct)]

            for col, val in enumerate(vals, 1):
                c = cel(ws, row, col, val, fill=bg)
                c.font = Font(name="Arial", size=9, color="000000")
                c.alignment = Alignment(
                    horizontal="left" if col == 2 else "center",
                    vertical="center",
                    indent=1 if col == 2 else 0,
                )
                ws.row_dimensions[row].height = 15

            c7 = ws.cell(row=row, column=7)
            c7.fill = mk_fill(fc)
            c7.font = Font(name="Arial", bold=True, size=9, color=ft)

        return row_start + 2 + len(locais)

    row_cur = 8

    top5_bom  = locais_sorted[:5]
    top5_ruim = list(reversed(locais_sorted[-5:]))

    row_cur = secao_ranking(row_cur, "🥇  TOP 5 MELHORES LOCAIS",
                             C_GREEN, top5_bom, medalhas=True)
    row_cur += 1

    row_cur = secao_ranking(row_cur,
                             "🚨  TOP 5 LOCAIS QUE PRECISAM DE ATENÇÃO",
                             C_RED, top5_ruim)
    row_cur += 1

    row_cur = secao_ranking(row_cur,
                             f"📋  RANKING COMPLETO  ({len(locais_sorted)} locais)",
                             C_DARK, locais_sorted)

    # Larguras
    widths = {"A": 6, "B": 36, "C": 9, "D": 9,
              "E": 10, "F": 12, "G": 12, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ─── ABA: Grade por Posto ─────────────────────────────────────────────────────

def build_grade(wb, dados: dict, start: str, end: str):
    ws = wb.create_sheet("Grade por Posto")
    ws.sheet_view.showGridLines = False

    records = dados["records"]
    ncols   = len(DIAS_ORDER) + 2

    merge(ws, 1, 1, 1, ncols)
    c = cel(ws, 1, 1,
            f"Grade de Programação por Posto  ·  {start} → {end}",
            fill=C_DARK)
    c.font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for col, txt in enumerate(["Local", "Posto / Modelo"] + DIAS_ORDER, 1):
        c = cel(ws, 2, col, txt, fill=C_DARK)
        c.font = mk_font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    grade = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))
    for r in records:
        if not r["avulsa"]:
            chave = f"{r['posto']} | {r['op_tipo']}"
            grade[r["local"]][chave][r["dia_semana"]].add(r["hora"])

    row_idx = 3
    for local in sorted(grade):
        for posto_key in sorted(grade[local]):
            dias_data = grade[local][posto_key]
            for col, val in [(1, local), (2, posto_key)]:
                c = cel(ws, row_idx, col, val, fill=C_LIGHT)
                c.font = Font(name="Arial", size=9, color="000000")
                c.alignment = Alignment(vertical="center",
                                        horizontal="left", indent=1)

            for col, dia in enumerate(DIAS_ORDER, start=3):
                horas = sorted(dias_data.get(dia, set()))
                val   = "\n".join(horas) if horas else ""
                c = cel(ws, row_idx, col,
                        val, fill=C_LGREEN if horas else None)
                c.font = Font(name="Arial", size=9,
                              color=C_GREEN if horas else "CCCCCC",
                              bold=bool(horas))
                c.alignment = Alignment(horizontal="center",
                                        vertical="center", wrap_text=True)

            h_max = max(
                (len(sorted(dias_data.get(d, set()))) for d in DIAS_ORDER),
                default=1,
            )
            ws.row_dimensions[row_idx].height = max(15 * h_max, 18)
            row_idx += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    for col in range(3, 3 + len(DIAS_ORDER)):
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.freeze_panes = "C3"



def slugify(text: str) -> str:
    """Converte texto para nome de arquivo seguro (sem acentos, sem espaços)."""
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", "_", text)
    return text.strip("_")


def build_arquivo_posto(dados: dict, posto_nome: str,
                        start: str, end: str) -> Workbook:
    """
    Arquivo por posto: tabela simples e plana com todas as atividades
    (ronda, limpeza, avulsa) de todos os locais que têm esse posto.
    AutoFilter em todas as colunas — fácil de filtrar/ordenar no Excel.
    """
    from openpyxl import Workbook as WB
    from openpyxl.worksheet.table import Table, TableStyleInfo

    posto_low  = posto_nome.strip().lower()
    recs_todos = sorted(
        [r for r in dados["records"]
         if r["posto"].strip().lower() == posto_low],
        key=lambda r: (r["local"], r["dt_sort"])
    )

    wb = WB()
    ws = wb.active
    ws.title = "Atividades"
    ws.sheet_view.showGridLines = False

    # ── Cabeçalho informativo (linhas 1-2) ──────────────────────────────────
    ok_g  = sum(1 for r in recs_todos if r["status_class"] == "Feita")
    par_g = sum(1 for r in recs_todos if r["status_class"] in ("Parcial","Parcial c/ Just."))
    nf_g  = sum(1 for r in recs_todos if r["status_class"] == "Não Feita")
    tot_g = len(recs_todos)
    pct_g = pct_eficiencia(ok_g, par_g, tot_g)
    n_loc = len({r["local"] for r in recs_todos})

    NCOLS = 12
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws.cell(row=1, column=1,
                value=f"  🔧  {posto_nome}   |   {n_loc} local(is)   |   "
                      f"Período: {start} → {end}   |   "
                      f"{tot_g} atividades   ·   "
                      f"✓ {ok_g} OK   ⚠ {par_g} Parcial   ✗ {nf_g} Não Feitas   "
                      f"·   {pct_g}% eficiência")
    c.font      = Font(name="Arial", bold=True, size=10, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 4   # espaçador

    # ── Cabeçalho da tabela (linha 3) ────────────────────────────────────────
    HEADERS = [
        "Local", "Posto", "Tipo", "Atividade",
        "Status", "Data", "Hora", "Turno",
        "Duração", "Pts OK", "Pts Total", "Justificativa",
    ]
    HDR_ROW = 3
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=10, color=C_WHITE)
        c.fill      = PatternFill("solid", fgColor=C_BLUE2)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[HDR_ROW].height = 20

    # ── Dados (a partir da linha 4) ──────────────────────────────────────────
    STATUS_FILL = {
        "Feita":            ("D4EDDA", "155724"),
        "Parcial":          ("FFF3CD", "856404"),
        "Parcial c/ Just.": ("FFF3CD", "856404"),
        "Não Feita":        ("F8D7DA", "721C24"),
        "Não iniciada":     ("E2E3E5", "383D41"),
    }

    for i, rec in enumerate(recs_todos):
        row = HDR_ROW + 1 + i
        bg  = "F8F9FA" if i % 2 == 0 else "FFFFFF"

        tipo = "Avulsa" if rec.get("avulsa") else (rec.get("op_tipo") or "Rotina").strip()
        if not tipo or tipo == "-":
            tipo = "Rotina"

        sf, st = STATUS_FILL.get(rec["status_class"], ("FFFFFF", "000000"))

        vals = [
            rec["local"],
            rec["posto"],
            tipo,
            rec["modelo"],
            rec["status_label"],
            rec["data"],
            rec["hora"],
            rec["turno"],
            rec["duracao"],
            rec["pontos_ok"],
            rec["pontos_total"],
            rec["justificativa"],
        ]
        CENTER_COLS = {5, 6, 7, 8, 9, 10, 11}   # Status, Data, Hora, Turno, Duração, Pts

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 5:   # Status — cor própria
                c.fill = PatternFill("solid", fgColor=sf)
                c.font = Font(name="Arial", bold=True, size=9, color=st)
            elif col == 3:  # Tipo — cor por categoria
                tipo_cor = {"Ronda": "1565C0", "Limpeza": "2E7D32",
                            "Avulsa": "E65100", "Rotina": "6A1B9A"}
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Arial", bold=True, size=9,
                              color=tipo_cor.get(tipo, "333333"))
            else:
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Arial", size=9, color="111111")
            c.alignment = Alignment(
                horizontal="center" if col in CENTER_COLS else "left",
                vertical="center",
                indent=1 if col in (1, 2, 4, 12) else 0,
            )
        ws.row_dimensions[row].height = 15

    # ── AutoFilter na tabela de dados ─────────────────────────────────────────
    last_row = HDR_ROW + len(recs_todos)
    ws.auto_filter.ref = (
        f"A{HDR_ROW}:{get_column_letter(NCOLS)}{max(last_row, HDR_ROW)}"
    )

    # ── Larguras ──────────────────────────────────────────────────────────────
    widths = [28, 20, 10, 36, 16, 12, 8, 16, 12, 8, 8, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{HDR_ROW + 1}"   # congela até linha de cabeçalho

    return wb


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    SEP = "═" * 56
    print(f"\n{SEP}")
    print("  FindMe  ·  Relatório de Atividades por Posto")
    print(SEP)

    cfg       = load_config()
    email     = cfg.get("email", "")
    password  = cfg.get("password", "")
    loc_uuids = cfg.get("locations", [])

    if not email or not password:
        print("❌  Configure email e password no config.json.")
        sys.exit(1)
    if not loc_uuids:
        print("❌  Informe ao menos 1 UUID em locations no config.json.")
        sys.exit(1)

    print()
    today     = date.today().isoformat()
    yesterday = (date.today() - timedelta(days=1)).isoformat()
    # Padrão diário: dia anterior completo + hoje até 07:00
    CUTOFF_HOUR = cfg.get("cutoff_hour", 7)   # configurável no config.json
    start = ask_date("Data início", yesterday)
    end   = ask_date("Data fim",    today)
    print(f"  ℹ   Período: {start} 00:00 → {end} {CUTOFF_HOUR:02d}:00 "
          f"(registros de {end} após {CUTOFF_HOUR:02d}:00 serão ignorados)\n"
          if start != end else "")
    print()

    print("  🔐  Autenticando...")
    token = login(email, password)
    print("  ✔   Token obtido.\n")

    print("  📍  Buscando locais...")
    all_locs = api_get(token, "/filters/locations")
    loc_map  = {
        l["uuid"]: l
        for l in
        (all_locs if isinstance(all_locs, list) else [])
        if isinstance(l, dict)
    }
    selected = [loc_map[uid] for uid in loc_uuids if uid in loc_map]
    print(f"  ✔   {len(selected)} local(is): "
          f"{', '.join(l['name'] for l in selected)}\n")

    print("  \U0001f4cb  Buscando atividades por local...")
    rows_raw   = []
    locs_vazios = []   # locais que retornaram 0 registros (para buscar histórico depois)
    for i, loc in enumerate(selected, 1):
        nome = loc["name"]
        print(f"  [{i:02d}/{len(selected)}] {nome}")
        filt_loc = {
            "hiddenInactive": True,
            "locations": [loc["uuid"]],
            "period": [start, end],
        }
        parcial, teve_504 = fetch_rotinas(token, filt_loc)

        # Se falhou com 504 E nao trouxe nada, tenta busca dia a dia automaticamente
        if teve_504 and len(parcial) == 0:
            print(f"  \U0001f501  504 persistente — retentando {nome} dia a dia...")
            parcial = fetch_rotinas_por_dia(token, filt_loc, start, end, nome)

        if len(parcial) == 0:
            locs_vazios.append(loc)

        rows_raw.extend(parcial)
        print(f"         → {len(parcial)} atividades (total: {len(rows_raw)})")

    print(f"\n  ✔   {len(rows_raw)} atividades no total.")

    # Aplica corte de horário: atividades do dia final só até CUTOFF_HOUR
    if start != end:
        cutoff_dt = datetime.strptime(end, "%Y-%m-%d").replace(
            hour=CUTOFF_HOUR, minute=0, second=0)
        antes = len(rows_raw)
        rows_raw = [
            r for r in rows_raw
            if not (
                parse_dt(r.get("to_be_started_at") or r.get("started_at") or "")
                is not None
                and parse_dt(r.get("to_be_started_at") or r.get("started_at") or "")
                >= cutoff_dt
                and parse_dt(r.get("to_be_started_at") or r.get("started_at") or "")
                .strftime("%Y-%m-%d") == end
            )
        ]
        cortados = antes - len(rows_raw)
        if cortados:
            print(f"  ✂   {cortados} registro(s) de {end} após "
                  f"{CUTOFF_HOUR:02d}:00 removido(s).")
    print()

    if not rows_raw:
        print("  ⚠   Nenhuma atividade encontrada no periodo.")
        sys.exit(0)

    print("  \U0001f50d  Processando dados...")
    dados = processar(rows_raw)
    por_local = dados["por_local"]

    print(f"       {len(por_local)} locais processados:")
    for nome, d in sorted(por_local.items(), key=lambda x: -x[1]["pct"]):
        barra = barra_progresso(d["pct"])
        print(f"       {nome[:35]:<35} {barra} {d['pct']:5.1f}%")

    # Injeta avulsas da pasta postos/ (um arquivo .json por local)
    import glob as _glob
    avulsas_cfg = []
    postos_files = sorted(_glob.glob("postos/*.json"))
    if postos_files:
        print(f"  📋  Carregando configuração de postos ({len(postos_files)} arquivo(s))...")
        for pf in postos_files:
            try:
                with open(pf, encoding="utf-8") as _f:
                    loc_data = json.load(_f)
                if "local" in loc_data and "postos" in loc_data:
                    avulsas_cfg.append(loc_data)
                    print(f"       ✔ {pf}")
                else:
                    print(f"  ⚠   {pf} sem chaves 'local'/'postos' — pulando.")
            except Exception as _e:
                print(f"  ⚠   Erro ao ler {pf}: {_e} — pulando.")
    else:
        print("  ℹ   Nenhum arquivo encontrado em postos/ — avulsas config desabilitada.")

    if avulsas_cfg:
        n_cfg = injetar_avulsas_config(dados, avulsas_cfg, start, end)
        print(f"  ✔   {n_cfg} avulsa(s) injetada(s) a partir de postos/.\n")

    # Busca avulsas perdidas
    print("  \U0001f514  Buscando avulsas perdidas (API missed-single-activities)...")
    todos_nomes = [l["name"] for l in selected]
    try:
        filt_all = {
            "hiddenInactive": True,
            "locations": [l["uuid"] for l in selected],
            "period": [start, end],
        }
        r_miss = requests.post(
            f"{BASE_DASHBOARD}/reports/missed-single-activities",
            headers={"Authorization": f"Bearer {token}",
                     "Content-Type": "application/json"},
            json=filt_all,
            params={"page": 1, "limit": 500},
            timeout=60,
        )
        if r_miss.status_code in (200, 201):
            miss_raw = r_miss.json()
            if isinstance(miss_raw, list):
                miss_list = miss_raw
            elif isinstance(miss_raw, dict):
                miss_list = (miss_raw.get("data") or miss_raw.get("rows")
                             or miss_raw.get("items") or [])
            else:
                miss_list = []
            n_inj = 0
            for row in miss_list:
                if not isinstance(row, dict):
                    continue
                local_nome = (row.get("location") or {}).get("name", "-")
                station    = row.get("station") or {}

                # Tenta todos os campos possíveis para data programada da avulsa
                dt_prog = None
                for campo_dt in ("to_be_started_at", "scheduled_at",
                                 "date", "programmed_at", "created_at"):
                    dt_prog = parse_dt(row.get(campo_dt))
                    if dt_prog:
                        break
                # Se ainda vazio, usa a data de inicio do periodo como referencia
                if dt_prog is None:
                    dt_prog = datetime.strptime(start, "%Y-%m-%d")

                # Respeita o corte de horario do dia final
                if (start != end
                        and dt_prog.strftime("%Y-%m-%d") == end
                        and dt_prog.hour >= CUTOFF_HOUR):
                    continue  # avulsa fora do periodo aceito

                rec = {
                    "local":         local_nome,
                    "regiao":        (row.get("region") or {}).get("name", "-"),
                    "cliente":       (row.get("client") or {}).get("name", "-"),
                    "posto":         station.get("name", "-"),
                    "op_tipo":       OP_TYPE.get(station.get("operation_type", 0), "N/C"),
                    "modelo":        (row.get("patrol") or {}).get("name", "-"),
                    "avulsa":        True,
                    # veio do endpoint missed-single-activities → é Perdida
                    # (antes: "Nao iniciada" sem acento, que nem casava com
                    # o mapa de cores e aparecia sem destaque)
                    "status_int":    5,
                    "status_label":  "Perdida",
                    "status_class":  "Não Feita",
                    "data":          data_fmt(dt_prog),
                    "hora":          hora_fmt(dt_prog),
                    "turno":         turno(dt_prog),
                    "hora_fim_prog": "-",
                    "iniciada":      "-",
                    "finalizada":    "-",
                    "duracao":       "-",
                    "dia_semana":    dia_semana(dt_prog),
                    "pontos_ok":     0,
                    "pontos_total":  0,
                    "justificativa": "-",
                    "dt_sort":       dt_prog or datetime.min,
                }
                dados["records"].append(rec)
                if local_nome not in dados["por_local"]:
                    dados["por_local"][local_nome] = {
                        "ok": 0, "parcial": 0, "nao_feita": 0,
                        "total": 0, "pct": 0.0, "records": []
                    }
                d2 = dados["por_local"][local_nome]
                d2["nao_feita"] += 1
                d2["total"] += 1
                d2["pct"] = pct_eficiencia(d2["ok"], d2["parcial"], d2["total"])
                d2["records"].append(rec)
                n_inj += 1
            print(f"  \u2714   {n_inj} avulsa(s) perdida(s) adicionada(s).\n")
        else:
            print(f"  \u26a0   missed-single-activities retornou {r_miss.status_code} - ignorando.\n")
    except Exception as e:
        print(f"  \u26a0   Erro ao buscar avulsas perdidas: {e} - ignorando.\n")

    # Busca modelos historicos para locais que ficaram completamente vazios
    modelos_historicos = {}
    if locs_vazios:
        print(f"  \U0001f50e  Buscando historico para {len(locs_vazios)} local(is) sem dados...")
        for loc in locs_vazios:
            nome_loc = loc["name"]
            print(f"      {nome_loc} — consultando 60 dias anteriores...")
            mods = buscar_modelos_historico(token, loc, start)
            if mods:
                modelos_historicos[nome_loc] = mods
                tipos_unicos = {m["op_tipo"] for m in mods}
                print(f"         \u2714  {len(mods)} modelo(s) encontrado(s): "
                      f"{', '.join(sorted(tipos_unicos))}")
            else:
                print(f"         \u26a0  Nenhum historico encontrado (local pode ser novo ou inativo)")
        print()

    # ── Pasta de saída relatorios/{start}_{end}/ ──────────────────────────────
    out_dir = os.path.join("relatorios", f"{start}_{end}")
    os.makedirs(out_dir, exist_ok=True)
    postos_dir = os.path.join(out_dir, "postos")
    os.makedirs(postos_dir, exist_ok=True)

    # ── Arquivo GERAL ──────────────────────────────────────────────────────────
    print("\n  📊  Gerando Excel GERAL...")
    wb = Workbook()
    wb.remove(wb.active)

    build_capa(wb, dados, start, end)
    build_atividades(wb, dados, start, end,
                     todos_locais=todos_nomes,
                     modelos_historicos=modelos_historicos)
    # Ranking removido do GERAL — redundante com a Capa Executiva e com o
    # PDF resumo. A função build_ranking segue definida caso queira reativar.
    build_grade(wb, dados, start, end)

    geral_name = f"GERAL_{start}_{end}.xlsx"
    geral_path = os.path.join(out_dir, geral_name)
    try:
        wb.save(geral_path)
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        geral_name = f"GERAL_{start}_{end}_{ts}.xlsx"
        geral_path = os.path.join(out_dir, geral_name)
        wb.save(geral_path)
        print(f"  ⚠   Arquivo anterior em uso — salvo como: {geral_name}")

    print(f"  ✅  GERAL salvo: {geral_path}")

    # ── Arquivos por POSTO ─────────────────────────────────────────────────────
    print("\n  🔧  Gerando arquivos por posto...")
    postos_gerados = 0
    postos_erros   = 0

    # Agrupa por nome de posto (un único arquivo por posto, unindo todos os locais)
    postos_unicos = sorted({
        r["posto"].strip()
        for r in dados["records"]
        if r["posto"] and r["posto"].strip() != "-"
    })

    for posto_nome in postos_unicos:
        # Quais locais têm esse posto?
        locais_do_posto = sorted({
            r["local"] for r in dados["records"]
            if r["posto"].strip().lower() == posto_nome.lower()
        })
        slug_posto = slugify(posto_nome)
        fname = f"{slug_posto}.xlsx"
        fpath = os.path.join(postos_dir, fname)
        try:
            wb_p = build_arquivo_posto(dados, posto_nome, start, end)
            try:
                wb_p.save(fpath)
            except PermissionError:
                ts = datetime.now().strftime("%H%M%S")
                fpath = fpath.replace(".xlsx", f"_{ts}.xlsx")
                wb_p.save(fpath)
            locais_str = ", ".join(locais_do_posto)
            print(f"       ✔  {posto_nome}  ({len(locais_do_posto)} local(is): {locais_str})")
            postos_gerados += 1
        except Exception as _ep:
            print(f"       ⚠  {posto_nome} — erro: {_ep}")
            postos_erros += 1

    print(f"\n{SEP}")
    print(f"  ✅  Relatórios gerados em:  relatorios/{start}_{end}/")
    print(f"      📄  GERAL:  {geral_name}")
    print(f"      🔧  Postos: {postos_gerados} arquivo(s) em postos/  {f'({postos_erros} erro(s))' if postos_erros else ''}")
    print(f"      Abas do GERAL:")
    print(f"        - Capa Executiva  -- KPIs + postos críticos/ótimos")
    print(f"        - Atividades      -- detalhe por local (turno, duração, status)")
    print(f"        - Ranking         -- top 5 melhores/piores + ranking completo")
    print(f"        - Grade por Posto -- visual: posto x dia x horário")
    print(f"      Abas de cada posto:")
    print(f"        - Resumo          -- KPIs do posto (ok, parcial, não feitas, eficiência)")
    print(f"        - Atividades      -- detalhe filtrado só para aquele posto")
    print(SEP + "\n")


if __name__ == "__main__":
    main()

