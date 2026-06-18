#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_aba_semana.py — Adiciona ao GERAL uma aba "Semana (7 dias)": um heatmap de
eficiência por LOCAL → POSTO ao longo dos últimos 7 dias, com média e tendência.

Fonte: a própria aba "Atividades" (já enriquecida) de cada GERAL diário em
relatorios/<dia>_<dia>/GERAL_<dia>_<dia>.xlsx. O dia atual é lido do workbook em
memória (ainda não salvo); os anteriores, dos arquivos. Eficiência por posto/dia =
(feitas + 0,5×parciais) / total, onde total inclui perdidas/não feitas/não
registradas — ou seja, % do que era esperado e foi efetivamente feito.

Uso como módulo (chamado pelo enriquecer_atividades.py antes de salvar):
    import gerar_aba_semana as gs
    gs.adicionar(wb, xlsx_path)
"""
from __future__ import annotations

import re
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ICONE_GRUPO = "📍"
ICONE_POSTO = "▸"

AZUL    = "1F4E79"
AZUL_CL = "DCE6F1"
VERDE   = "C6EFCE"; VERDE_TX = "1E6B3C"
AMAR    = "FFEB9C"; AMAR_TX  = "7F6000"
VERM    = "FFC7CE"; VERM_TX  = "9C0006"
CINZA   = "EDEAE2"; CINZA_TX = "7F7F7F"
BRANCO  = "FFFFFF"
DIAS_PT = ["seg", "ter", "qua", "qui", "sex", "sáb", "dom"]

_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _stats_atividades(ws) -> dict:
    """Lê uma aba 'Atividades' (enriquecida) -> {(local, posto): {feita,parcial,nao,total}}."""
    stats, local, posto = {}, None, None
    for row in ws.iter_rows(min_row=2, max_col=3):
        v1 = str(row[0].value or "")
        if ICONE_GRUPO in v1:
            local = re.split(r"\s{2,}", v1.replace(ICONE_GRUPO, "").strip())[0].strip()
            posto = None
            continue
        if ICONE_POSTO in v1:
            posto = re.split(r"\s{2,}", v1.replace(ICONE_POSTO, "").strip())[0].strip()
            continue
        if local is None or posto is None:
            continue
        modelo = str(row[1].value or "").strip()
        if not modelo:
            continue
        status = str(row[2].value or "").strip().lower()
        s = stats.setdefault((local, posto),
                             {"feita": 0, "parcial": 0, "nao": 0, "total": 0})
        s["total"] += 1
        if status == "feita":
            s["feita"] += 1
        elif status == "parcial":
            s["parcial"] += 1
        else:                       # perdida, não feita, não registrada
            s["nao"] += 1
    return stats


def _stats_arquivo(caminho: Path) -> dict:
    try:
        wb = load_workbook(caminho, read_only=True, data_only=True)
    except Exception:
        return {}
    if "Atividades" not in wb.sheetnames:
        wb.close()
        return {}
    st = _stats_atividades(wb["Atividades"])
    wb.close()
    return st


def _efic(s: dict):
    t = s.get("total", 0)
    if not t:
        return None
    return round((s["feita"] + 0.5 * s["parcial"]) / t * 100)


def _cor(efic):
    if efic is None:
        return CINZA, CINZA_TX
    if efic >= 90:
        return VERDE, VERDE_TX
    if efic >= 70:
        return VERDE, VERDE_TX
    if efic >= 40:
        return AMAR, AMAR_TX
    return VERM, VERM_TX


def _glob_geral(relatorios_dir: Path, d: str):
    """Acha o GERAL canônico do dia d (relatorios/d_d/GERAL_d_d*.xlsx)."""
    pasta = relatorios_dir / f"{d}_{d}"
    cano = pasta / f"GERAL_{d}_{d}.xlsx"
    if cano.exists():
        return cano
    cands = sorted(pasta.glob(f"GERAL_{d}_{d}*.xlsx"))
    return cands[-1] if cands else None


def adicionar(wb, xlsx_path, n_dias: int = 7) -> bool:
    """Insere a aba 'Semana (7 dias)' no workbook `wb` (o GERAL do dia atual).
    `xlsx_path` aponta pra relatorios/<dia>_<dia>/GERAL_...xlsx — usado pra achar
    a pasta e os dias anteriores. O dia atual é lido do próprio `wb`."""
    xlsx_path = Path(xlsx_path)
    pasta = xlsx_path.parent
    relatorios_dir = pasta.parent
    try:
        dia_op = pasta.name.split("_")[0]
        d0 = datetime.strptime(dia_op, "%Y-%m-%d")
    except (ValueError, IndexError):
        return False

    dias = [(d0 - timedelta(days=i)).strftime("%Y-%m-%d")
            for i in range(n_dias - 1, -1, -1)]   # antigo -> recente

    # stats por dia
    por_dia = {}
    for d in dias:
        if d == dia_op:
            por_dia[d] = _stats_atividades(wb["Atividades"]) if "Atividades" in wb.sheetnames else {}
        else:
            cam = _glob_geral(relatorios_dir, d)
            por_dia[d] = _stats_arquivo(cam) if cam else {}

    # universo de (local, posto), preservando a ordem do dia mais recente
    ordem, vistos = [], set()
    for d in reversed(dias):
        for k in por_dia[d].keys():
            if k not in vistos:
                vistos.add(k)
                ordem.append(k)
    # agrupa por local mantendo ordem de aparição
    locais_ordem, postos_por_local = [], {}
    for (loc, posto) in ordem:
        if loc not in postos_por_local:
            postos_por_local[loc] = []
            locais_ordem.append(loc)
        postos_por_local[loc].append(posto)

    # ── monta a aba ───────────────────────────────────────────────────────────
    nome_aba = "Semana (7 dias)"
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]
    ws = wb.create_sheet(nome_aba)
    ws.sheet_view.showGridLines = False

    headers = ["Local / Posto"] + [
        f"{datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m')}\n{DIAS_PT[datetime.strptime(d, '%Y-%m-%d').weekday()]}"
        for d in dias
    ] + ["Média", "Tend."]
    ncols = len(headers)

    # título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1,
                value=f"Análise da semana — eficiência por posto  "
                      f"({datetime.strptime(dias[0], '%Y-%m-%d').strftime('%d/%m')} a "
                      f"{d0.strftime('%d/%m/%Y')})")
    t.font = Font(name="Arial", bold=True, color=BRANCO, size=12)
    t.fill = PatternFill("solid", start_color=AZUL)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    # cabeçalho
    r = 2
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color=BRANCO, size=9)
        c.fill = PatternFill("solid", start_color=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
    ws.row_dimensions[r].height = 28
    r += 1

    def _tend(vals):
        v = [x for x in vals if x is not None]
        if len(v) < 2:
            return "—", CINZA_TX
        d = v[-1] - v[0]
        if d >= 10:
            return "▲", VERDE_TX
        if d <= -10:
            return "▼", VERM_TX
        return "▬", CINZA_TX

    for loc in locais_ordem:
        # linha-cabeçalho do local (média do local na semana)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cl = ws.cell(row=r, column=1, value=f"{ICONE_GRUPO}  {loc}")
        cl.font = Font(name="Arial", bold=True, color=AZUL, size=10)
        cl.fill = PatternFill("solid", start_color=AZUL_CL)
        cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        r += 1

        for posto in postos_por_local[loc]:
            efs = []
            for d in dias:
                s = por_dia[d].get((loc, posto))
                efs.append(_efic(s) if s else None)
            # nome do posto
            cp = ws.cell(row=r, column=1, value=f"   {ICONE_POSTO} {posto}")
            cp.font = Font(name="Arial", size=9)
            cp.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cp.border = _BORDER
            # dias
            for j, ef in enumerate(efs):
                c = ws.cell(row=r, column=2 + j,
                            value=(f"{ef}%" if ef is not None else "—"))
                bg, fg = _cor(ef)
                c.fill = PatternFill("solid", start_color=bg)
                c.font = Font(name="Arial", size=9, bold=False, color=fg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = _BORDER
            # média
            v = [x for x in efs if x is not None]
            media = round(sum(v) / len(v)) if v else None
            cm = ws.cell(row=r, column=2 + len(dias),
                         value=(f"{media}%" if media is not None else "—"))
            bg, fg = _cor(media)
            cm.fill = PatternFill("solid", start_color=bg)
            cm.font = Font(name="Arial", size=9, bold=True, color=fg)
            cm.alignment = Alignment(horizontal="center", vertical="center")
            cm.border = _BORDER
            # tendência
            simb, cor = _tend(efs)
            ct = ws.cell(row=r, column=3 + len(dias), value=simb)
            ct.font = Font(name="Arial", size=11, bold=True, color=cor)
            ct.alignment = Alignment(horizontal="center", vertical="center")
            ct.border = _BORDER
            r += 1

    # legenda
    r += 1
    leg = ws.cell(row=r, column=1,
                  value="Eficiência = (feitas + ½ parciais) ÷ total esperado.  "
                        "Verde ≥70% · Amarelo 40–70% · Vermelho <40% · — sem dado.  "
                        "Tendência compara o 1º e o último dia com dado (▲ melhora · ▼ piora).")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    leg.font = Font(name="Arial", size=8, italic=True, color=CINZA_TX)
    leg.alignment = Alignment(horizontal="left", vertical="center")

    # larguras + freeze
    ws.column_dimensions["A"].width = 42
    for j in range(len(dias)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 9
    ws.column_dimensions[get_column_letter(2 + len(dias))].width = 9
    ws.column_dimensions[get_column_letter(3 + len(dias))].width = 7
    ws.freeze_panes = "B3"
    return True
