#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_locais_enderecos.py — Gera um Excel com os locais monitorados e o endereço
de cada um, via API do FindMe.

Fonte: `/filters/locations` (nome/região) + `/v3/settings/locations/{uuid}`
(endereço, latitude/longitude) da `production.api.findme.id` (Bearer token).
NÃO inclui as senhas de acesso que esse endpoint também devolve.

Uso:
    python gerar_locais_enderecos.py [saida.xlsx]   # default: LOCAIS_ENDERECOS.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import findme_programacao as fp

AZUL = "1F4E79"
ZEBRA = "F2F6FB"
BRANCO = "FFFFFF"
LINK = "0563C1"
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
V3 = "https://production.api.findme.id/v3/settings/locations"


def coletar(cfg: dict) -> list:
    token = fp.login(cfg["email"], cfg["password"])
    locs = fp.api_get(token, "/filters/locations")
    meta = {l["uuid"]: l for l in locs if isinstance(l, dict) and l.get("uuid")}
    linhas = []
    for uuid in cfg.get("locations", []):
        m = meta.get(uuid, {})
        nome = m.get("client_name") or m.get("name") or uuid
        regiao = m.get("region_name", "") or ""
        end, lat, lng = "—", None, None
        try:
            d = requests.get(f"{V3}/{uuid}",
                             headers={"Authorization": f"Bearer {token}"},
                             timeout=30).json()
            end = (d.get("address") or "").strip() or "—"
            lat, lng = d.get("latitude"), d.get("longitude")
        except Exception as e:
            print(f"  ⚠ {nome[:30]}: {type(e).__name__}", file=sys.stderr)
        linhas.append({"nome": nome, "endereco": end, "lat": lat,
                       "lng": lng, "regiao": regiao})
    linhas.sort(key=lambda x: x["nome"].upper())
    return linhas


def gerar(saida) -> int:
    saida = Path(saida)
    cfg = fp.load_config("config.json")
    linhas = coletar(cfg)

    wb = Workbook()
    ws = wb.active
    ws.title = "Locais"
    ws.sheet_view.showGridLines = False

    headers = ["#", "Local", "Endereço", "Coordenadas", "Mapa", "Região"]
    widths = [4, 40, 58, 22, 12, 20]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color=BRANCO, size=10)
        c.fill = PatternFill("solid", start_color=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for i, l in enumerate(linhas, start=1):
        r = i + 1
        bg = ZEBRA if i % 2 == 0 else BRANCO
        coord = (f"{l['lat']:.5f}, {l['lng']:.5f}"
                 if l["lat"] and l["lng"] else "—")
        vals = [str(i), l["nome"], l["endereco"], coord, "", l["regiao"]]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = PatternFill("solid", start_color=bg)
            c.border = _BORDER
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(
                horizontal="center" if ci in (1, 4) else "left",
                vertical="center", wrap_text=(ci == 3), indent=(1 if ci in (2, 3, 6) else 0))
        # link do Google Maps na coluna 5
        if l["lat"] and l["lng"]:
            cm = ws.cell(row=r, column=5, value="abrir mapa")
            cm.hyperlink = f"https://maps.google.com/?q={l['lat']},{l['lng']}"
            cm.font = Font(name="Arial", size=9, color=LINK, underline="single")
            cm.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 28

    ws.cell(row=len(linhas) + 3, column=1,
            value=f"{len(linhas)} locais · fonte: FindMe (v3/settings/locations) · "
                  "coordenadas com link pro Google Maps").font = Font(
        name="Arial", size=8, italic=True, color="7F7F7F")

    wb.save(saida)
    return len(linhas)


def main():
    saida = sys.argv[1] if len(sys.argv) > 1 else "LOCAIS_ENDERECOS.xlsx"
    n = gerar(saida)
    print(f"OK: {n} locais em {saida}")


if __name__ == "__main__":
    main()
