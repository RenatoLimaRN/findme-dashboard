#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_pdf_whatsapp.py — Gera UM PDF único com o resumo por local, no estilo das
mensagens de WhatsApp (substitui os .txt soltos por local).

Lê a aba "Atividades" do GERAL enriquecido e monta, para cada local, um cartão
com: cabeçalho (nome + feitas/total + eficiência), e por posto (▸) a lista de
CADA execução em ordem cronológica do plantão (06h→05h), além das atividades
esperadas que o sistema nem registrou.

Marcadores de status (coloridos, fonte DejaVu pra renderizar/copiar certo):
    ✓ feita   ◐ parcial   ✗ não feita / perdida   □ não registrada

Uso como módulo:
    import gerar_pdf_whatsapp as gw
    gw.gerar("GERAL_2026-06-13.xlsx", "2026-06-13", Path("WHATSAPP_2026-06-13.pdf"))

Uso por linha de comando:
    python gerar_pdf_whatsapp.py <geral.xlsx> <YYYY-MM-DD> <saida.pdf>
"""
from __future__ import annotations

import re
import sys
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (KeepTogether, Paragraph, SimpleDocTemplate,
                                Spacer)

# ─── Paleta (mesma do resumo) ────────────────────────────────────────────────
AZUL_H   = "#1F4E79"
VERDE_H  = "#1E6B3C"
AMR_H    = "#7F6000"
VERM_H   = "#C0504D"
CINZA_H  = "#5F5E5A"

ICONE_GRUPO = "📍"
ICONE_POSTO = "▸"

# Conjuntos de símbolos por fonte. DejaVu renderiza os glifos bonitos; o
# fallback Helvetica (WinAnsi) usa ASCII pra não sair lixo se faltar matplotlib.
_SIMBOLOS = {
    True: {  # DejaVu
        "mark": {"feita": ("✓", VERDE_H), "parcial": ("◐", AMR_H),
                 "perdida": ("✗", VERM_H), "não feita": ("✗", VERM_H),
                 "nao feita": ("✗", VERM_H)},
        "padrao": ("•", CINZA_H), "bullet": "•", "naoreg": "□",
        "legenda": "✓ feita · ◐ parcial · ✗ não feita · □ não registrada",
    },
    False: {  # Helvetica / WinAnsi
        "mark": {"feita": ("OK", VERDE_H), "parcial": ("~", AMR_H),
                 "perdida": ("X", VERM_H), "não feita": ("X", VERM_H),
                 "nao feita": ("X", VERM_H)},
        "padrao": ("-", CINZA_H), "bullet": "-", "naoreg": "[ ]",
        "legenda": "OK feita · ~ parcial · X nao feita · [ ] nao registrada",
    },
}


def _registrar_fontes() -> tuple[str, str, bool]:
    """Registra DejaVuSans (acentos + ✓◐✗□▸) via matplotlib. Fallback: Helvetica.
    Retorna (fonte, fonte_bold, tem_dejavu)."""
    try:
        import matplotlib
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.pdfmetrics import registerFontFamily
        from reportlab.pdfbase.ttfonts import TTFont
        base = Path(matplotlib.__file__).parent / "mpl-data" / "fonts" / "ttf"
        reg, bold = base / "DejaVuSans.ttf", base / "DejaVuSans-Bold.ttf"
        if reg.exists() and bold.exists():
            pdfmetrics.registerFont(TTFont("DejaVu", str(reg)))
            pdfmetrics.registerFont(TTFont("DejaVu-Bold", str(bold)))
            registerFontFamily("DejaVu", normal="DejaVu", bold="DejaVu-Bold",
                               italic="DejaVu", boldItalic="DejaVu-Bold")
            return "DejaVu", "DejaVu-Bold", True
    except Exception:
        pass
    return "Helvetica", "Helvetica-Bold", False


def _coletar(xlsx_path) -> list:
    """Lê a aba Atividades -> [{nome, postos:[{posto, linhas:[(quando,modelo,status)]}]}]."""
    wb = load_workbook(xlsx_path, read_only=True)
    if "Atividades" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Atividades"]
    locais, atual, posto_atual = [], None, None
    for row in ws.iter_rows(min_row=2, max_col=3):
        v1 = str(row[0].value or "")
        if ICONE_GRUPO in v1:
            nome = re.split(r"\s{2,}", v1.replace(ICONE_GRUPO, "").strip())[0].strip()
            atual = {"nome": nome, "postos": []}
            posto_atual = None
            locais.append(atual)
            continue
        if ICONE_POSTO in v1:
            nome_p = re.split(r"\s{2,}", v1.replace(ICONE_POSTO, "").strip())[0].strip()
            posto_atual = {"posto": nome_p, "linhas": []}
            if atual is not None:
                atual["postos"].append(posto_atual)
            continue
        if atual is None or posto_atual is None:
            continue
        quando = v1.strip()
        modelo = str(row[1].value or "").strip()
        status = str(row[2].value or "").strip()
        if modelo:
            posto_atual["linhas"].append((quando, modelo, status))
    wb.close()
    return locais


def _faixa_cor(pct: float) -> str:
    if pct >= 90:
        return VERDE_H
    if pct >= 70:
        return VERDE_H
    if pct >= 40:
        return AMR_H
    return VERM_H


def gerar(xlsx_path, data_alvo: str, saida) -> Path:
    saida = Path(saida)
    saida.parent.mkdir(parents=True, exist_ok=True)

    fonte, fonte_b, tem_dejavu = _registrar_fontes()
    sym = _SIMBOLOS[tem_dejavu]
    locais = _coletar(xlsx_path)

    d0 = datetime.fromisoformat(data_alvo)
    d1 = d0 + timedelta(days=1)
    plantao = f"{d0.strftime('%d/%m')} 06h → {d1.strftime('%d/%m')} 05h"

    styles = getSampleStyleSheet()
    st_titulo = ParagraphStyle("t", parent=styles["Normal"], fontName=fonte_b,
                               fontSize=16, textColor=colors.HexColor(AZUL_H),
                               leading=19)
    st_sub = ParagraphStyle("s", parent=styles["Normal"], fontName=fonte,
                            fontSize=9.5, textColor=colors.HexColor(CINZA_H),
                            leading=12, spaceAfter=2)
    st_local = ParagraphStyle("l", parent=styles["Normal"], fontName=fonte_b,
                              fontSize=12, textColor=colors.HexColor(AZUL_H),
                              leading=15, spaceBefore=8, spaceAfter=1)
    st_posto = ParagraphStyle("p", parent=styles["Normal"], fontName=fonte_b,
                              fontSize=9.5, textColor=colors.HexColor(CINZA_H),
                              leading=13, spaceBefore=4, spaceAfter=1)
    st_linha = ParagraphStyle("li", parent=styles["Normal"], fontName=fonte,
                              fontSize=9, leading=13, alignment=TA_LEFT,
                              textColor=colors.HexColor("#222222"), leftIndent=8)
    st_esp = ParagraphStyle("e", parent=styles["Normal"], fontName=fonte,
                            fontSize=8.5, leading=12, leftIndent=8,
                            textColor=colors.HexColor(CINZA_H))
    st_rodape = ParagraphStyle("r", parent=styles["Normal"], fontName=fonte,
                               fontSize=7.5, textColor=colors.HexColor(CINZA_H),
                               leading=11, alignment=2)

    doc = SimpleDocTemplate(
        str(saida), pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=f"WhatsApp FindMe {d0.strftime('%d/%m/%Y')}",
    )
    elems = [
        Paragraph("FindMe — mensagens por local", st_titulo),
        Paragraph(f"Plantão {plantao} · texto pronto por local", st_sub),
        Spacer(1, 6),
    ]

    n_locais = 0
    for loc in locais:
        if not loc.get("postos"):
            continue

        todas = [x for p in loc["postos"] for x in p["linhas"]]
        reais_tot = [x for x in todas if "/" in x[0]]
        if not reais_tot and not any("/" not in x[0] for x in todas):
            continue
        feitas_tot = sum(1 for _, _, s in reais_tot if s.lower() == "feita")
        pct = (100.0 * feitas_tot / len(reais_tot)) if reais_tot else 0.0
        cor_efic = _faixa_cor(pct)

        bloco = [Paragraph(
            f'{escape(loc["nome"])}  '
            f'<font size=9 color="{cor_efic}">— {feitas_tot} de {len(reais_tot)} '
            f'feitas · {pct:.0f}%</font>', st_local)]

        for p in loc["postos"]:
            reais = [x for x in p["linhas"] if "/" in x[0]]
            esperadas = [x for x in p["linhas"] if "/" not in x[0]]
            if not reais and not esperadas:
                continue
            bloco.append(Paragraph(f"{ICONE_POSTO} {escape(p['posto'])}", st_posto))
            for quando, modelo, status in reais:
                simb, cor = sym["mark"].get(status.lower(), sym["padrao"])
                bloco.append(Paragraph(
                    f'<font color="{cor}"><b>{escape(simb)}</b></font> '
                    f'<b>{escape(quando)}</b>  {escape(modelo)}', st_linha))
            if esperadas:
                bloco.append(Paragraph(
                    f'<font color="{CINZA_H}">{escape(sym["naoreg"])} não registradas '
                    f'({len(esperadas)}):</font>', st_esp))
                for _, modelo, _ in esperadas:
                    bloco.append(Paragraph(f"{escape(sym['bullet'])} {escape(modelo)}", st_esp))

        bloco.append(Spacer(1, 8))
        elems.append(KeepTogether(bloco))
        n_locais += 1

    if n_locais == 0:
        elems.append(Paragraph("Sem locais com atividades neste plantão.", st_sub))

    gerado = datetime.now().strftime("%d/%m/%Y %H:%M")
    elems.append(Spacer(1, 4))
    elems.append(Paragraph(
        f"Gerado automaticamente em {gerado} · {sym['legenda']}", st_rodape))

    doc.build(elems)
    return saida


def main():
    if len(sys.argv) < 4:
        print("uso: python gerar_pdf_whatsapp.py <geral.xlsx> <YYYY-MM-DD> <saida.pdf>",
              file=sys.stderr)
        sys.exit(1)
    out = gerar(sys.argv[1], sys.argv[2], Path(sys.argv[3]))
    print(f"PDF gerado: {out}")


if __name__ == "__main__":
    main()
