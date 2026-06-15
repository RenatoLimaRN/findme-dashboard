#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_textos_whatsapp.py — Gera um texto pronto (pra copiar/colar no WhatsApp)
por local, a partir da aba "Atividades" do relatório GERAL enriquecido.

Cada local vira um arquivo .txt em <saida_dir>/<slug>.txt. O texto lista CADA
execução (não agrupa) com data+hora, em ordem cronológica do plantão (06h→05h),
e no fim as atividades esperadas que o sistema nem registrou.

Uso como módulo:
    import gerar_textos_whatsapp as gw
    gw.gerar("GERAL_2026-06-13.xlsx", "2026-06-13", "saida/whatsapp")

Uso por linha de comando:
    python gerar_textos_whatsapp.py <geral.xlsx> <YYYY-MM-DD> <saida_dir>
"""
from __future__ import annotations

import re
import sys
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

EMOJI = {
    "feita": "✅", "parcial": "🟡", "perdida": "❌",
    "não feita": "❌", "nao feita": "❌",
}
ICONE_GRUPO = "📍"


def _slug(nome: str) -> str:
    s = unicodedata.normalize("NFD", nome or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"^\s*CONDOMINIO\s+", "", s, flags=re.I).strip()
    return re.sub(r"[^A-Za-z0-9]+", "_", s.lower()).strip("_") or "local"


ICONE_POSTO = "▸"


def _coletar(xlsx_path) -> list:
    """Lê a aba Atividades -> [{nome, postos:[{posto, linhas:[(quando,modelo,status)]}]}]."""
    wb = load_workbook(xlsx_path, read_only=True)
    if "Atividades" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Atividades"]
    locais, atual, posto_atual = [], None, None
    for row in ws.iter_rows(min_row=2, max_col=3):
        v1 = str(row[0].value or "")
        if ICONE_GRUPO in v1:
            nome = re.split(r"\s{2,}", v1.replace(ICONE_GRUPO, "").strip())[0].strip()
            atual = {"nome": nome, "postos": []}
            posto_atual = None
            locais.append(atual)
            continue
        if ICONE_POSTO in v1:
            nome_p = re.split(r"\s{2,}", v1.replace(ICONE_POSTO, "").strip())[0].strip()
            posto_atual = {"posto": nome_p, "linhas": []}
            if atual is not None:
                atual["postos"].append(posto_atual)
            continue
        if atual is None or posto_atual is None:
            continue
        quando = v1.strip()
        modelo = str(row[1].value or "").strip()
        status = str(row[2].value or "").strip()
        if modelo:
            posto_atual["linhas"].append((quando, modelo, status))
    wb.close()
    return locais


def _texto_local(loc: dict, data_alvo: str) -> str:
    d0 = datetime.fromisoformat(data_alvo)
    d1 = d0 + timedelta(days=1)
    plantao = f"{d0.strftime('%d/%m')} 06h → {d1.strftime('%d/%m')} 05h"

    todas = [x for p in loc["postos"] for x in p["linhas"]]
    reais_tot = [x for x in todas if "/" in x[0]]
    feitas_tot = sum(1 for _, _, s in reais_tot if s.lower() == "feita")

    L = [f"*FINDME — {loc['nome']}*",
         f"_Plantão {plantao}_",
         "",
         f"{feitas_tot} de {len(reais_tot)} atividades feitas"]

    for p in loc["postos"]:
        reais = [x for x in p["linhas"] if "/" in x[0]]
        esperadas = [x for x in p["linhas"] if "/" not in x[0]]
        if not reais and not esperadas:
            continue
        L.append("")
        L.append(f"*▸ {p['posto']}*")
        for quando, modelo, status in reais:
            e = EMOJI.get(status.lower(), "•")
            L.append(f"{e} {quando}  {modelo}")
        if esperadas:
            L.append(f"⬜ _não registradas ({len(esperadas)}):_")
            for _, modelo, _ in esperadas:
                L.append(f"• {modelo}")

    return "\n".join(L)


def gerar(xlsx_path, data_alvo: str, saida_dir) -> list:
    saida_dir = Path(saida_dir)
    saida_dir.mkdir(parents=True, exist_ok=True)
    locais = _coletar(xlsx_path)
    gerados = []
    for loc in locais:
        if not loc.get("postos"):
            continue
        texto = _texto_local(loc, data_alvo)
        caminho = saida_dir / f"{_slug(loc['nome'])}.txt"
        caminho.write_text(texto + "\n", encoding="utf-8")
        gerados.append(caminho)
    return gerados


def main():
    if len(sys.argv) < 4:
        print("uso: python gerar_textos_whatsapp.py <geral.xlsx> <YYYY-MM-DD> <saida_dir>",
              file=sys.stderr)
        sys.exit(1)
    gerados = gerar(sys.argv[1], sys.argv[2], sys.argv[3])
    print(f"{len(gerados)} texto(s) gerado(s) em {sys.argv[3]}")


if __name__ == "__main__":
    main()
